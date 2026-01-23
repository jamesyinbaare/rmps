"""
Service for exporting candidate processed results to Excel.
"""

import io
import re
from typing import Any

import pandas as pd
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Candidate,
    Exam,
    ExamRegistration,
    ExamSeries,
    ExamSubject,
    ExamType,
    Programme,
    School,
    Subject,
    SubjectRegistration,
    SubjectScore,
    SubjectType,
)
from app.utils.score_utils import calculate_grade


def sanitize_filename_part(text: str) -> str:
    """
    Sanitize a string to be safe for use in filenames.
    Replaces spaces with underscores and removes invalid characters.
    """
    if not text:
        return ""
    # Replace spaces with underscores
    text = text.replace(" ", "_")
    # Remove invalid filename characters: / \ : * ? " < > |
    text = re.sub(r'[<>:"/\\|?*]', '', text)
    # Remove multiple consecutive underscores
    text = re.sub(r'_+', '_', text)
    # Remove leading/trailing underscores
    text = text.strip('_')
    return text


async def generate_export_filename(
    session: AsyncSession,
    exam_id: int | None = None,
    exam_type: ExamType | None = None,
    series: ExamSeries | None = None,
    year: int | None = None,
    subject_type: SubjectType | None = None,
    programme_id: int | None = None,
    subject_id: int | None = None,
) -> str:
    """
    Generate a descriptive filename for the export based on filters.

    Format: {exam_year}_{exam_series}_{exam_type}_{additional_options}_scores.xlsx

    Additional options (in order):
    - If subject_type is CORE: add "CORE"
    - If subject_type is ELECTIVE: add "ELECTIVE"
    - If programme_id is provided: add programme name
    - If subject_id is provided: add subject name

    Args:
        session: Database session
        exam_id: Exam ID (preferred for getting exam details)
        exam_type: Exam type (fallback if exam_id not provided)
        series: Exam series (fallback if exam_id not provided)
        year: Exam year (fallback if exam_id not provided)
        subject_type: Subject type filter (CORE or ELECTIVE)
        programme_id: Programme ID filter
        subject_id: Subject ID filter

    Returns:
        Sanitized filename string
    """
    parts = []

    # Get exam details
    exam_year = None
    exam_series_str = None
    exam_type_str = None

    if exam_id is not None:
        # Fetch exam from database
        exam_stmt = select(Exam).where(Exam.id == exam_id)
        exam_result = await session.execute(exam_stmt)
        exam = exam_result.scalar_one_or_none()
        if exam:
            exam_year = exam.year
            exam_series_str = exam.series.value if hasattr(exam.series, 'value') else str(exam.series)
            exam_type_str = exam.exam_type.value if hasattr(exam.exam_type, 'value') else str(exam.exam_type)
    else:
        # Use provided values as fallback
        exam_year = year
        if series:
            exam_series_str = series.value if hasattr(series, 'value') else str(series)
        if exam_type:
            exam_type_str = exam_type.value if hasattr(exam_type, 'value') else str(exam_type)

    # Build base parts
    if exam_year:
        parts.append(str(exam_year))
    if exam_series_str:
        parts.append(exam_series_str)
    if exam_type_str:
        parts.append(sanitize_filename_part(exam_type_str))

    # Add additional options
    if subject_type == SubjectType.CORE:
        parts.append("CORE")
    elif subject_type == SubjectType.ELECTIVE:
        parts.append("ELECTIVE")

    # Add programme name if provided
    if programme_id is not None:
        programme_stmt = select(Programme).where(Programme.id == programme_id)
        programme_result = await session.execute(programme_stmt)
        programme = programme_result.scalar_one_or_none()
        if programme:
            parts.append(sanitize_filename_part(programme.name))
        else:
            parts.append("Unknown_Programme")

    # Add subject name if provided
    if subject_id is not None:
        subject_stmt = select(Subject).where(Subject.id == subject_id)
        subject_result = await session.execute(subject_stmt)
        subject = subject_result.scalar_one_or_none()
        if subject:
            parts.append(sanitize_filename_part(subject.name))
        else:
            parts.append("Unknown_Subject")

    # Add suffix
    parts.append("scores")

    # Join parts with underscores
    filename = "_".join(parts)

    # If no parts were added, use fallback
    if not filename or filename == "scores":
        filename = "candidate_results_export"

    # Limit total length (conservative limit for filesystem compatibility)
    if len(filename) > 200:
        filename = filename[:200]

    # Add extension
    final_filename = f"{filename}.xlsx"
    return final_filename


# Field definitions with human-readable labels
EXPORT_FIELDS = {
    # Candidate fields
    "candidate_name": "Candidate Name",
    "candidate_index_number": "Index Number",
    # School fields
    "school_name": "School Name",
    "school_code": "School Code",
    # Exam fields
    "exam_name": "Exam Name",
    "exam_type": "Exam Type",
    "exam_year": "Exam Year",
    "exam_series": "Exam Series",
    # Programme fields
    "programme_name": "Programme Name",
    "programme_code": "Programme Code",
    # Subject fields
    "subject_name": "Subject Name",
    "subject_code": "Subject Code",
    "subject_series": "Subject Series",
    # Raw scores
    "obj_raw_score": "Objectives Raw Score",
    "essay_raw_score": "Essay Raw Score",
    "pract_raw_score": "Practical Raw Score",
    # Normalized scores
    "obj_normalized": "Objectives Normalized",
    "essay_normalized": "Essay Normalized",
    "pract_normalized": "Practical Normalized",
    # Results
    "total_score": "Total Score",
    "grade": "Grade",
    # Document IDs
    "obj_document_id": "Objectives Document ID",
    "essay_document_id": "Essay Document ID",
    "pract_document_id": "Practical Document ID",
    # Metadata
    "created_at": "Created At",
    "updated_at": "Updated At",
}


async def generate_results_export(
    session: AsyncSession,
    exam_id: int | None = None,
    exam_type: ExamType | None = None,
    series: ExamSeries | None = None,
    year: int | None = None,
    school_id: int | None = None,
    programme_id: int | None = None,
    subject_id: int | None = None,
    document_id: str | None = None,
    fields: list[str] | None = None,
    subject_type: SubjectType | None = None,
    export_format: str = "standard",
    test_type: str | None = None,
    subject_ids: list[int] | None = None,
) -> bytes:
    """
    Generate Excel export for candidate processed results.

    Args:
        session: Database session
        exam_id: Filter by exam ID
        exam_type: Filter by exam type
        series: Filter by exam series
        year: Filter by exam year
        school_id: Filter by school ID
        programme_id: Filter by programme ID
        subject_id: Filter by subject ID
        document_id: Filter by document ID (matches obj/essay/pract document IDs)
        fields: List of field names to include in export. If None, includes all fields.
        subject_type: Filter by subject type (CORE or ELECTIVE). If ELECTIVE, programme_id is required.
        export_format: "standard" or "multi_subject" - determines export format
        test_type: "obj" or "essay" - required for multi_subject format
        subject_ids: List of subject IDs - used for multi_subject format

    Returns:
        Bytes of Excel file

    Raises:
        ValueError: If invalid fields are provided, no results found, or invalid filter combination
    """
    # Route to multi-subject export if requested
    if export_format == "multi_subject":
        return await generate_multi_subject_export(
            session=session,
            exam_id=exam_id,
            exam_type=exam_type,
            series=series,
            year=year,
            school_id=school_id,
            programme_id=programme_id,
            test_type=test_type or "obj",
            subject_ids=subject_ids,
            subject_type=subject_type,
            fields=fields,
        )

    # Continue with standard export format
    # Validate fields if provided
    if fields is not None:
        invalid_fields = [f for f in fields if f not in EXPORT_FIELDS]
        if invalid_fields:
            raise ValueError(f"Invalid fields: {', '.join(invalid_fields)}")

    # Validate filter combinations
    if subject_type is not None and subject_id is not None:
        raise ValueError("subject_type and subject_id cannot both be specified")

    if subject_type == SubjectType.ELECTIVE and programme_id is None:
        raise ValueError("programme_id is required when subject_type is ELECTIVE")

    # Build query - same as get_candidates_for_manual_entry
    base_stmt = (
        select(
            Candidate,
            SubjectRegistration,
            SubjectScore,
            ExamRegistration,
            Exam,
            ExamSubject,
            Subject,
            Programme,
            School,
        )
        .join(SubjectScore, SubjectScore.subject_registration_id == SubjectRegistration.id)
        .join(ExamRegistration, SubjectRegistration.exam_registration_id == ExamRegistration.id)
        .join(Candidate, ExamRegistration.candidate_id == Candidate.id)
        .join(School, Candidate.school_id == School.id)
        .join(Exam, ExamRegistration.exam_id == Exam.id)
        .join(ExamSubject, SubjectRegistration.exam_subject_id == ExamSubject.id)
        .join(Subject, ExamSubject.subject_id == Subject.id)
        .outerjoin(Programme, Candidate.programme_id == Programme.id)
    )

    # Apply filters
    if exam_id is not None:
        base_stmt = base_stmt.where(Exam.id == exam_id)
    else:
        if exam_type is not None:
            base_stmt = base_stmt.where(Exam.exam_type == exam_type)
        if series is not None:
            base_stmt = base_stmt.where(Exam.series == series)
        if year is not None:
            base_stmt = base_stmt.where(Exam.year == year)
    if school_id is not None:
        base_stmt = base_stmt.where(Candidate.school_id == school_id)
    if programme_id is not None:
        base_stmt = base_stmt.where(Candidate.programme_id == programme_id)
    if subject_id is not None:
        base_stmt = base_stmt.where(Subject.id == subject_id)
    if subject_type is not None:
        base_stmt = base_stmt.where(Subject.subject_type == subject_type)
    if document_id is not None:
        base_stmt = base_stmt.where(
            or_(
                SubjectScore.obj_document_id == document_id,
                SubjectScore.essay_document_id == document_id,
                SubjectScore.pract_document_id == document_id,
            )
        )

    # Execute query - get all results (no pagination for export)
    stmt = base_stmt.order_by(Candidate.index_number, Subject.code)
    result = await session.execute(stmt)
    rows = result.all()

    if not rows:
        raise ValueError("No results found matching the specified filters")

    # Helper function to build a row from candidate data
    def build_row(candidate, subject_reg, subject_score, _exam_reg, exam, exam_subject, subject, programme, school, fields_to_export):
        grade = calculate_grade(
            subject_score.total_score,
            exam_subject.grade_ranges_json,
            subject_score=subject_score,
            exam_subject=exam_subject,
        )

        row_data: dict[str, Any] = {}

        if "candidate_name" in fields_to_export:
            row_data["Candidate Name"] = candidate.name
        if "candidate_index_number" in fields_to_export:
            row_data["Index Number"] = candidate.index_number
        if "school_name" in fields_to_export:
            row_data["School Name"] = school.name
        if "school_code" in fields_to_export:
            row_data["School Code"] = school.code
        if "exam_name" in fields_to_export:
            row_data["Exam Name"] = exam.exam_type.value
        if "exam_type" in fields_to_export:
            row_data["Exam Type"] = exam.exam_type.value
        if "exam_year" in fields_to_export:
            row_data["Exam Year"] = exam.year
        if "exam_series" in fields_to_export:
            row_data["Exam Series"] = exam.series.value
        if "programme_name" in fields_to_export:
            row_data["Programme Name"] = programme.name if programme else None
        if "programme_code" in fields_to_export:
            row_data["Programme Code"] = programme.code if programme else None
        if "subject_name" in fields_to_export:
            row_data["Subject Name"] = subject.name
        if "subject_code" in fields_to_export:
            row_data["Subject Code"] = subject.code
        if "subject_series" in fields_to_export:
            row_data["Subject Series"] = subject_reg.series
        if "obj_raw_score" in fields_to_export:
            row_data["Objectives Raw Score"] = subject_score.obj_raw_score
        if "essay_raw_score" in fields_to_export:
            row_data["Essay Raw Score"] = subject_score.essay_raw_score
        if "pract_raw_score" in fields_to_export:
            row_data["Practical Raw Score"] = subject_score.pract_raw_score
        if "obj_normalized" in fields_to_export:
            row_data["Objectives Normalized"] = subject_score.obj_normalized
        if "essay_normalized" in fields_to_export:
            row_data["Essay Normalized"] = subject_score.essay_normalized
        if "pract_normalized" in fields_to_export:
            row_data["Practical Normalized"] = subject_score.pract_normalized
        if "total_score" in fields_to_export:
            row_data["Total Score"] = subject_score.total_score
        if "grade" in fields_to_export:
            row_data["Grade"] = grade.value if grade else None
        if "obj_document_id" in fields_to_export:
            row_data["Objectives Document ID"] = subject_score.obj_document_id
        if "essay_document_id" in fields_to_export:
            row_data["Essay Document ID"] = subject_score.essay_document_id
        if "pract_document_id" in fields_to_export:
            row_data["Practical Document ID"] = subject_score.pract_document_id
        if "created_at" in fields_to_export:
            row_data["Created At"] = subject_score.created_at.strftime("%Y-%m-%d %H:%M:%S") if subject_score.created_at else None
        if "updated_at" in fields_to_export:
            row_data["Updated At"] = subject_score.updated_at.strftime("%Y-%m-%d %H:%M:%S") if subject_score.updated_at else None

        return row_data

    # Determine which fields to export
    fields_to_export = fields if fields is not None else list(EXPORT_FIELDS.keys())

    # Filter and group rows based on subject_type or subject_id
    if subject_type == SubjectType.CORE:
        # Filter to only core subjects (already filtered in query), then group by subject (each on separate sheet)
        if not rows:
            raise ValueError("No core subjects found matching the specified filters")

        # Group by subject
        grouped_data: dict[tuple[int, str, str], list] = {}  # (subject_id, subject_code, subject_name) -> rows
        for row in rows:
            subject = row[6]  # Subject is at index 6
            key = (subject.id, subject.code, subject.name)
            if key not in grouped_data:
                grouped_data[key] = []
            grouped_data[key].append(row)

    elif subject_type == SubjectType.ELECTIVE:
        # Filter to only elective subjects for the specified programme, then group by subject (each on separate sheet)
        if not programme_id:
            raise ValueError("programme_id is required when subject_type is ELECTIVE")

        # Get elective subjects for this programme
        from app.models import programme_subjects
        programme_elective_stmt = select(Subject.id).select_from(
            programme_subjects.join(Subject, programme_subjects.c.subject_id == Subject.id)
        ).where(
            programme_subjects.c.programme_id == programme_id,
            Subject.subject_type == SubjectType.ELECTIVE
        )
        elective_result = await session.execute(programme_elective_stmt)
        elective_subject_ids = {r[0] for r in elective_result.all()}

        # Filter rows to only include elective subjects for this programme
        # Note: subject_type filter is already applied in the query, but we need to ensure
        # the subjects are actually electives for this programme and candidates are in the programme
        filtered_rows = []
        for row in rows:
            candidate, subject_reg, subject_score, _exam_reg, exam, exam_subject, subject, programme, school = row
            # Double-check: subject must be ELECTIVE type and in the programme's elective list
            if subject.subject_type == SubjectType.ELECTIVE and subject.id in elective_subject_ids:
                # Ensure candidate is in the specified programme (programme filter should already be applied)
                if programme and programme.id == programme_id:
                    filtered_rows.append(row)

        if not filtered_rows:
            raise ValueError("No elective subjects found for the specified programme")

        # Group by subject
        grouped_data: dict[tuple[int, str, str], list] = {}  # (subject_id, subject_code, subject_name) -> rows
        for row in filtered_rows:
            subject = row[6]  # Subject is at index 6
            key = (subject.id, subject.code, subject.name)
            if key not in grouped_data:
                grouped_data[key] = []
            grouped_data[key].append(row)

    elif subject_id is not None:
        # Single subject export - all in one sheet
        # Get subject info from first row
        if rows:
            subject = rows[0][6]  # Subject is at index 6
            grouped_data = {(subject.id, subject.code, subject.name): rows}
        else:
            raise ValueError("No results found for the specified subject")

    else:
        # No subject_type or subject_id specified - export all in one sheet
        grouped_data = {(0, "Candidate Results", "Candidate Results"): rows}

    # Generate Excel file with multiple sheets
    output = io.BytesIO()
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        used_sheet_names: set[str] = set()
        sheet_counter: dict[str, int] = {}

        for (group_subject_id, subject_code, subject_name), group_rows in sorted(grouped_data.items()):
            # Build rows for this group
            export_rows = []
            for row in group_rows:
                candidate, subject_reg, subject_score, _exam_reg, exam, exam_subject, subject, programme, school = row
                row_data = build_row(candidate, subject_reg, subject_score, _exam_reg, exam, exam_subject, subject, programme, school, fields_to_export)
                export_rows.append(row_data)

            # Create DataFrame for this sheet
            df = pd.DataFrame(export_rows)

            # Generate sheet name (max 31 chars for Excel, must be unique)
            # Use format: "CODE - Name" or just "CODE" if name makes it too long
            if subject_type == SubjectType.CORE or subject_type == SubjectType.ELECTIVE or (subject_id is not None and group_subject_id != 0):
                # Try to use "CODE - Name" format, but truncate if needed
                combined_name = f"{subject_code} - {subject_name}"
                if len(combined_name) <= 29:  # Leave room for suffix if needed
                    base_sheet_name = combined_name
                else:
                    # If too long, try just code and truncated name
                    max_code_len = len(subject_code) + 3  # " - " separator
                    if max_code_len < 29:
                        truncated_name = subject_name[:29 - max_code_len]
                        base_sheet_name = f"{subject_code} - {truncated_name}"
                    else:
                        # If code itself is too long, just use code
                        base_sheet_name = subject_code[:29]
            else:
                base_sheet_name = "Candidate Results"[:31]

            # Ensure unique sheet name
            sheet_name = base_sheet_name
            if sheet_name in used_sheet_names:
                counter = sheet_counter.get(base_sheet_name, 1)
                sheet_counter[base_sheet_name] = counter + 1
                sheet_name = f"{base_sheet_name[:27]}_{counter}"[:31]

            used_sheet_names.add(sheet_name)

            # Write to Excel
            df.to_excel(writer, index=False, sheet_name=sheet_name)

            # Auto-size columns
            worksheet = writer.sheets[sheet_name]
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(str(col))
                ) if len(df) > 0 else len(str(col))
                col_letter = get_column_letter(idx)
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output.seek(0)
    return output.getvalue()


async def generate_multi_subject_export(
    session: AsyncSession,
    exam_id: int | None = None,
    exam_type: ExamType | None = None,
    series: ExamSeries | None = None,
    year: int | None = None,
    school_id: int | None = None,
    programme_id: int | None = None,
    test_type: str = "obj",
    subject_ids: list[int] | None = None,
    subject_type: SubjectType | None = None,
    fields: list[str] | None = None,
) -> bytes:
    """
    Generate Excel export with multiple subjects on the same sheet.

    Format: Each row is a candidate, columns are candidate info fields + subject codes.
    For each subject column, shows the raw score for the selected test_type, or "N/A" if not registered.

    Args:
        session: Database session
        exam_id: Filter by exam ID
        exam_type: Filter by exam type
        series: Filter by exam series
        year: Filter by exam year
        school_id: Filter by school ID
        programme_id: Filter by programme ID
        test_type: "obj" or "essay" - which raw score to export
        subject_ids: List of subject IDs to include (mutually exclusive with subject_type)
        subject_type: Subject type filter (CORE or ELECTIVE) - mutually exclusive with subject_ids
        fields: List of candidate info fields to include

    Returns:
        Bytes of Excel file

    Raises:
        ValueError: If invalid configuration or no results found
    """
    # Validate fields if provided
    if fields is not None:
        # Filter out subject-specific fields for multi-subject format
        allowed_fields = [
            "candidate_name", "candidate_index_number", "school_name", "school_code",
            "exam_name", "exam_type", "exam_year", "exam_series",
            "programme_name", "programme_code"
        ]
        invalid_fields = [f for f in fields if f not in allowed_fields]
        if invalid_fields:
            raise ValueError(f"Invalid fields for multi-subject format: {', '.join(invalid_fields)}")

    # Determine which fields to export
    fields_to_export = fields if fields is not None else ["candidate_name", "candidate_index_number"]

    # Step 1: Get all candidates for the exam (with filters)
    candidate_stmt = (
        select(Candidate, ExamRegistration, Exam, School, Programme)
        .join(ExamRegistration, ExamRegistration.candidate_id == Candidate.id)
        .join(Exam, ExamRegistration.exam_id == Exam.id)
        .join(School, Candidate.school_id == School.id)
        .outerjoin(Programme, Candidate.programme_id == Programme.id)
    )

    # Apply filters
    if exam_id is not None:
        candidate_stmt = candidate_stmt.where(Exam.id == exam_id)
    else:
        if exam_type is not None:
            candidate_stmt = candidate_stmt.where(Exam.exam_type == exam_type)
        if series is not None:
            candidate_stmt = candidate_stmt.where(Exam.series == series)
        if year is not None:
            candidate_stmt = candidate_stmt.where(Exam.year == year)
    if school_id is not None:
        candidate_stmt = candidate_stmt.where(Candidate.school_id == school_id)
    if programme_id is not None:
        candidate_stmt = candidate_stmt.where(Candidate.programme_id == programme_id)

    candidate_stmt = candidate_stmt.order_by(Candidate.index_number)
    candidate_result = await session.execute(candidate_stmt)
    candidate_rows = candidate_result.all()

    if not candidate_rows:
        raise ValueError("No candidates found matching the specified filters")

    # Step 2: Determine which subjects to include
    selected_subject_ids = set()

    if subject_ids is not None:
        # Use provided subject IDs
        selected_subject_ids = set(subject_ids)

        # Validate that these subjects exist in the exam
        exam_subject_stmt = (
            select(ExamSubject.subject_id)
            .join(Exam, ExamSubject.exam_id == Exam.id)
        )
        if exam_id is not None:
            exam_subject_stmt = exam_subject_stmt.where(Exam.id == exam_id)
        else:
            if exam_type is not None:
                exam_subject_stmt = exam_subject_stmt.where(Exam.exam_type == exam_type)
            if series is not None:
                exam_subject_stmt = exam_subject_stmt.where(Exam.series == series)
            if year is not None:
                exam_subject_stmt = exam_subject_stmt.where(Exam.year == year)

        exam_subject_result = await session.execute(exam_subject_stmt)
        exam_subject_ids = {row[0] for row in exam_subject_result.all()}

        # Filter to only subjects that exist in the exam
        selected_subject_ids = selected_subject_ids & exam_subject_ids

        if not selected_subject_ids:
            raise ValueError("None of the specified subject IDs exist in the exam")

    elif subject_type is not None:
        # Get subjects by type
        subject_stmt = select(Subject.id).join(ExamSubject, Subject.id == ExamSubject.subject_id).join(Exam, ExamSubject.exam_id == Exam.id)

        if exam_id is not None:
            subject_stmt = subject_stmt.where(Exam.id == exam_id)
        else:
            if exam_type is not None:
                subject_stmt = subject_stmt.where(Exam.exam_type == exam_type)
            if series is not None:
                subject_stmt = subject_stmt.where(Exam.series == series)
            if year is not None:
                subject_stmt = subject_stmt.where(Exam.year == year)

        subject_stmt = subject_stmt.where(Subject.subject_type == subject_type)

        if subject_type == SubjectType.ELECTIVE and programme_id is not None:
            # For electives, also filter by programme
            from app.models import programme_subjects
            subject_stmt = subject_stmt.join(
                programme_subjects, Subject.id == programme_subjects.c.subject_id
            ).where(programme_subjects.c.programme_id == programme_id)

        subject_result = await session.execute(subject_stmt)
        selected_subject_ids = {row[0] for row in subject_result.all()}

        if not selected_subject_ids:
            raise ValueError(f"No {subject_type.value} subjects found for the specified exam")
    else:
        raise ValueError("Either subject_ids or subject_type must be provided")

    # Step 3: Get subject codes for the selected subjects
    subject_code_stmt = select(Subject.id, Subject.code).where(Subject.id.in_(selected_subject_ids)).order_by(Subject.code)
    subject_code_result = await session.execute(subject_code_stmt)
    subject_codes_map = {row[0]: row[1] for row in subject_code_result.all()}
    subject_codes_sorted = sorted(subject_codes_map.values())

    # Step 4: Get all subject registrations and scores for these candidates
    # Get exam registration IDs for the candidates
    exam_reg_ids = {row[1].id for row in candidate_rows}  # ExamRegistration is at index 1

    # Create a map from exam_registration_id to candidate_id
    exam_reg_to_candidate: dict[int, int] = {}
    for candidate, exam_reg, exam, school, programme in candidate_rows:
        exam_reg_to_candidate[exam_reg.id] = candidate.id

    # Get subject registrations (to know which subjects candidates registered for)
    subject_reg_stmt = (
        select(SubjectRegistration, SubjectScore, Subject, ExamSubject, ExamRegistration)
        .join(ExamSubject, SubjectRegistration.exam_subject_id == ExamSubject.id)
        .join(Subject, ExamSubject.subject_id == Subject.id)
        .join(ExamRegistration, SubjectRegistration.exam_registration_id == ExamRegistration.id)
        .outerjoin(SubjectScore, SubjectRegistration.id == SubjectScore.subject_registration_id)
        .where(
            SubjectRegistration.exam_registration_id.in_(exam_reg_ids),
            Subject.id.in_(selected_subject_ids)
        )
    )

    subject_reg_result = await session.execute(subject_reg_stmt)
    subject_reg_rows = subject_reg_result.all()

    # Step 5: Build pivot structure
    # Map: candidate_id -> {subject_code: value}
    # Value can be:
    #   - raw_score string if registered and has score
    #   - "" (empty string) if registered but no score
    #   - "N/A" if not registered
    candidate_scores: dict[int, dict[str, str]] = {}

    # Initialize all candidates with "N/A" (not registered)
    for candidate, exam_reg, exam, school, programme in candidate_rows:
        candidate_scores[candidate.id] = {code: "N/A" for code in subject_codes_sorted}

    # Fill in scores from subject registrations
    for subject_reg, subject_score, subject, exam_subject, exam_reg in subject_reg_rows:
        candidate_id = exam_reg.candidate_id
        subject_code = subject.code

        if subject_code in subject_codes_sorted:
            # Candidate registered for this subject
            # Get the raw score for the selected test_type
            if subject_score:
                if test_type == "obj":
                    raw_score = subject_score.obj_raw_score
                elif test_type == "essay":
                    raw_score = subject_score.essay_raw_score
                else:
                    raw_score = None
            else:
                raw_score = None

            if candidate_id in candidate_scores:
                # If registered but no score, leave blank (empty string)
                # If registered and has score, use the score
                candidate_scores[candidate_id][subject_code] = raw_score if raw_score is not None else ""

    # Step 6: Build Excel rows
    export_rows = []
    for candidate, exam_reg, exam, school, programme in candidate_rows:
        row_data: dict[str, Any] = {}

        # Add candidate info fields
        if "candidate_name" in fields_to_export:
            row_data["Candidate Name"] = candidate.name
        if "candidate_index_number" in fields_to_export:
            row_data["Index Number"] = candidate.index_number
        if "school_name" in fields_to_export:
            row_data["School Name"] = school.name
        if "school_code" in fields_to_export:
            row_data["School Code"] = school.code
        if "exam_name" in fields_to_export:
            row_data["Exam Name"] = exam.exam_type.value
        if "exam_type" in fields_to_export:
            row_data["Exam Type"] = exam.exam_type.value
        if "exam_year" in fields_to_export:
            row_data["Exam Year"] = exam.year
        if "exam_series" in fields_to_export:
            row_data["Exam Series"] = exam.series.value
        if "programme_name" in fields_to_export:
            row_data["Programme Name"] = programme.name if programme else None
        if "programme_code" in fields_to_export:
            row_data["Programme Code"] = programme.code if programme else None

        # Add subject scores
        candidate_id = candidate.id
        if candidate_id in candidate_scores:
            for subject_code in subject_codes_sorted:
                value = candidate_scores[candidate_id][subject_code]
                # value is already set correctly:
                # - raw_score string if registered and has score
                # - "" (empty string) if registered but no score
                # - "N/A" if not registered
                row_data[subject_code] = value
        else:
            # Should not happen, but handle gracefully
            for subject_code in subject_codes_sorted:
                row_data[subject_code] = "N/A"

        export_rows.append(row_data)

    # Step 7: Generate Excel
    output = io.BytesIO()
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(export_rows)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Generate sheet name
        sheet_name = "Multi_Subject_Scores"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Auto-size columns
        worksheet = writer.sheets[sheet_name]
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            ) if len(df) > 0 else len(str(col))
            col_letter = get_column_letter(idx)
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output.seek(0)
    return output.getvalue()
