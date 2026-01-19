"""Utility for generating Excel template files for bulk uploads."""

import io
from typing import Any

import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.styles import PatternFill
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models import ExamSubject, Subject, SubjectType


def generate_programme_template() -> bytes:
    """
    Generate Excel template for programme upload.

    Returns:
        Bytes of Excel file
    """
    data = {
        "code": ["PROG01", "PROG02"],
        "name": ["Example Programme 1", "Example Programme 2"],
        "exam_type": ["Certificate II Examinations", "Advance"],
    }
    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Programmes")
    output.seek(0)
    return output.getvalue()


def generate_subject_template() -> bytes:
    """
    Generate Excel template for subject upload.

    Returns:
        Bytes of Excel file
    """
    data = {
        "code": ["301", "302", "701", "NVTI-001"],
        "original_code": ["C30-1-01", "C30-1-02", "C701", "NVTI-001"],
        "name": ["Mathematics", "English", "Science", "NVTI Subject"],
        "subject_type": ["CORE", "CORE", "ELECTIVE", "ELECTIVE"],
        "exam_type": ["Certificate II Examinations", "Certificate II Examinations", "Advance", "Diploma"],
        "programme_type": ["CERT2", "CERT2", "CERT2", "NVTI"],
        "programme_code": ["PROG01", "PROG01", "", ""],
    }
    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Subjects")
    output.seek(0)
    return output.getvalue()


def generate_candidate_template() -> bytes:
    """
    Generate Excel template for candidate upload.

    Returns:
        Bytes of Excel file
    """
    data = {
        "school_code": ["SCH001", "SCH001"],
        "programme_code": ["PROG01", "PROG01"],
        "name": ["John Doe", "Jane Smith"],
        "index_number": ["1234567890", "0987654321"],
        "subject_codes": ["MAT,ENG,SCI", "MAT,ENG"],
    }
    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Candidates")
    output.seek(0)
    return output.getvalue()


def generate_school_template() -> bytes:
    """
    Generate Excel template for school upload.

    Returns:
        Bytes of Excel file
    """
    data = {
        "code": ["SCH001", "SCH002"],
        "name": ["Example School 1", "Example School 2"],
        "region": ["Greater Accra Region", "Ashanti Region"],
        "zone": ["A", "B"],
        "school_type": ["PUBLIC", "PRIVATE"],
        "programme_codes": ["PROG01,PROG02", "PROG01"],
    }
    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Schools")
    output.seek(0)
    return output.getvalue()


async def generate_exam_subject_template(
    session: AsyncSession, exam_id: int, subject_type: SubjectType | None = None
) -> bytes:
    """
    Generate Excel template for exam subject upload.

    Args:
        session: Database session
        exam_id: Exam ID to generate template for
        subject_type: Optional filter by subject type (CORE or ELECTIVE)

    Returns:
        Bytes of Excel file
    """
    # Query ExamSubject records joined with Subject for the specified exam_id
    exam_subject_stmt = (
        select(ExamSubject, Subject)
        .join(Subject, ExamSubject.subject_id == Subject.id)
        .where(ExamSubject.exam_id == exam_id)
    )

    # Filter by subject_type if provided
    if subject_type is not None:
        exam_subject_stmt = exam_subject_stmt.where(Subject.subject_type == subject_type)

    exam_subject_stmt = exam_subject_stmt.order_by(Subject.code)

    result = await session.execute(exam_subject_stmt)
    exam_subjects_data = result.all()

    if not exam_subjects_data:
        # Return empty template with headers if no subjects
        data = {
            "original_code": [],
            "subject_name": [],
            "obj_pct": [],
            "essay_pct": [],
            "pract_pct": [],
            "obj_max_score": [],
            "essay_max_score": [],
        }
        df = pd.DataFrame(data)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="ExamSubjects")
            # Access workbook to add validation and formulas (0 rows in this case)
            workbook = writer.book
            worksheet = workbook.active
            _add_exam_subject_validation_and_formulas(worksheet, 0)
        output.seek(0)
        return output.getvalue()

    # Build DataFrame with subject data
    rows = []
    for exam_subject, subject in exam_subjects_data:
        rows.append(
            {
                "original_code": subject.original_code,
                "subject_name": subject.name,
                "obj_pct": "",  # Leave blank for user input
                "essay_pct": "",  # Leave blank for user input
                "pract_pct": "",  # Leave blank for user input
                "obj_max_score": "",  # Leave blank for user input
                "essay_max_score": "",  # Leave blank for user input
            }
        )

    df = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="ExamSubjects")
        # Access workbook to add validation and formulas
        workbook = writer.book
        worksheet = workbook.active
        _add_exam_subject_validation_and_formulas(worksheet, len(rows))

    output.seek(0)
    return output.getvalue()


def _add_exam_subject_validation_and_formulas(worksheet, num_rows: int) -> None:
    """
    Add data validation and formula column to exam subject template worksheet.

    Args:
        worksheet: openpyxl worksheet object
        num_rows: Number of data rows (excluding header)
    """
    # Column indices (1-indexed in openpyxl, but 0-indexed for column letters)
    # A = original_code, B = subject_name, C = obj_pct, D = essay_pct, E = pract_pct,
    # F = obj_max_score, G = essay_max_score, H = Total % Check (formula column)
    COL_ORIGINAL_CODE = "A"
    COL_SUBJECT_NAME = "B"
    COL_OBJ_PCT = "C"
    COL_ESSAY_PCT = "D"
    COL_PRACT_PCT = "E"
    COL_OBJ_MAX_SCORE = "F"
    COL_ESSAY_MAX_SCORE = "G"
    COL_TOTAL_PCT_CHECK = "H"

    # Add header for formula column
    worksheet[f"{COL_TOTAL_PCT_CHECK}1"] = "Total % Check"

    # Create data validation for percentage columns (>= 0, allow blank)
    pct_validation = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
    )
    pct_validation.error = "Percentage must be greater than or equal to 0"
    pct_validation.errorTitle = "Invalid Input"

    # Create data validation for max score columns (> 0, allow blank)
    score_validation = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
    )
    score_validation.error = "Max score must be greater than 0"
    score_validation.errorTitle = "Invalid Input"

    # Apply validations to rows (starting from row 2, row 1 is header)
    for row_num in range(2, num_rows + 2):
        # Add validation to percentage columns
        pct_validation.add(f"{COL_OBJ_PCT}{row_num}")
        pct_validation.add(f"{COL_ESSAY_PCT}{row_num}")
        pct_validation.add(f"{COL_PRACT_PCT}{row_num}")

        # Add validation to score columns
        score_validation.add(f"{COL_OBJ_MAX_SCORE}{row_num}")
        score_validation.add(f"{COL_ESSAY_MAX_SCORE}{row_num}")

        # Add formula to check that percentages sum to 100
        # Formula: =IF(AND(ISBLANK(C{row}),ISBLANK(D{row}),ISBLANK(E{row})), "", IF(SUM(C{row},D{row},E{row})=100, "OK", "Should equal 100"))
        formula = f'=IF(AND(ISBLANK({COL_OBJ_PCT}{row_num}),ISBLANK({COL_ESSAY_PCT}{row_num}),ISBLANK({COL_PRACT_PCT}{row_num})), "", IF(SUM({COL_OBJ_PCT}{row_num},{COL_ESSAY_PCT}{row_num},{COL_PRACT_PCT}{row_num})=100, "OK", "Should equal 100"))'
        worksheet[f"{COL_TOTAL_PCT_CHECK}{row_num}"] = formula

    # Add validations to worksheet
    worksheet.add_data_validation(pct_validation)
    worksheet.add_data_validation(score_validation)

    # Add conditional formatting
    if num_rows > 0:
        _add_conditional_formatting(worksheet, num_rows)


def _add_conditional_formatting(worksheet, num_rows: int) -> None:
    """
    Add conditional formatting to exam subject template worksheet.

    Args:
        worksheet: openpyxl worksheet object
        num_rows: Number of data rows (excluding header)
    """
    # Column indices
    COL_OBJ_PCT = "C"
    COL_ESSAY_PCT = "D"
    COL_PRACT_PCT = "E"
    COL_OBJ_MAX_SCORE = "F"
    COL_ESSAY_MAX_SCORE = "G"
    COL_TOTAL_PCT_CHECK = "H"

    # Define colors
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Format percentage columns: Green if >= 0 and valid, red if < 0 or invalid
    # Green: value is >= 0 (or blank which is treated as 0)
    # Red: value is < 0
    for row_num in range(2, num_rows + 2):
        # Percentage columns - green if >= 0 (including blank), red if < 0
        for col in [COL_OBJ_PCT, COL_ESSAY_PCT, COL_PRACT_PCT]:
            # Green: value >= 0 or is blank
            green_rule = FormulaRule(
                formula=[f'OR(ISBLANK({col}{row_num}), {col}{row_num}>=0)'],
                fill=green_fill,
            )
            worksheet.conditional_formatting.add(f"{col}{row_num}", green_rule)

            # Red: value < 0
            red_rule = CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)
            worksheet.conditional_formatting.add(f"{col}{row_num}", red_rule)

        # Max score columns - green if > 0, red if <= 0 (including blank treated as invalid for max_scores)
        for col in [COL_OBJ_MAX_SCORE, COL_ESSAY_MAX_SCORE]:
            # Green: value > 0
            green_rule = CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill)
            worksheet.conditional_formatting.add(f"{col}{row_num}", green_rule)

            # Red: value <= 0 or blank (max_scores must be positive)
            red_rule = FormulaRule(
                formula=[f'OR(ISBLANK({col}{row_num}), {col}{row_num}<=0)'],
                fill=red_fill,
            )
            worksheet.conditional_formatting.add(f"{col}{row_num}", red_rule)

        # Total % Check column - green if "OK", red if "Should equal 100"
        green_rule = FormulaRule(
            formula=[f'{COL_TOTAL_PCT_CHECK}{row_num}="OK"'],
            fill=green_fill,
        )
        worksheet.conditional_formatting.add(f"{COL_TOTAL_PCT_CHECK}{row_num}", green_rule)

        red_rule = FormulaRule(
            formula=[f'{COL_TOTAL_PCT_CHECK}{row_num}="Should equal 100"'],
            fill=red_fill,
        )
        worksheet.conditional_formatting.add(f"{COL_TOTAL_PCT_CHECK}{row_num}", red_rule)
