"""Super-admin: list and export exam centre officials across centres."""

import io
import re
import zipfile
from collections import defaultdict
from collections.abc import Sequence
from datetime import datetime
from typing import Literal, cast
from uuid import UUID

from fastapi import APIRouter, HTTPException, Query, Response, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, select
from sqlalchemy.orm import selectinload

from app.dependencies.auth import SuperAdminDep
from app.dependencies.database import DBSessionDep
from app.models import ExamCentreOfficial, Examination, ExamOfficialDesignation, School
from app.schemas.admin_exam_official import AdminExamCentreOfficialListResponse, AdminExamCentreOfficialRow

router = APIRouter(prefix="/admin/exam-centre-officials", tags=["admin-exam-officials"])

_MAX_LIST = 500
_DEFAULT_LIST = 100

_HEADER_LABELS = [
    "Centre code",
    "Centre name",
    "Full name",
    "Designation",
    "Bank",
    "Branch",
    "Bank code",
    "Account",
    "Days",
    "Phone",
]


def _designation_str(des: object) -> str:
    if isinstance(des, ExamOfficialDesignation):
        return des.value
    return str(des)


def _examination_label(ex: Examination) -> str:
    parts = [str(ex.year)]
    if ex.exam_series and str(ex.exam_series).strip():
        parts.append(str(ex.exam_series).strip())
    parts.append(str(ex.exam_type).strip())
    return " ".join(parts)


def _safe_filename_part(s: str) -> str:
    t = re.sub(r"[^\w\-]+", "_", s.strip(), flags=re.UNICODE).strip("_")
    return (t or "export")[:80]


def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _style_header_row(ws: object, row: int, ncols: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor="E8EDF4")
    font = Font(bold=True, size=11)
    b = _thin_border()
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = b
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_data_row(ws: object, row: int, ncols: int) -> None:
    b = _thin_border()
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = b
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def _set_col_widths(ws: object, widths: Sequence[float]) -> None:
    from openpyxl.utils import get_column_letter

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _data_values(off: ExamCentreOfficial, school: School) -> tuple[str | int, ...]:
    bb = off.bank_branch
    return (
        school.code,
        school.name,
        off.full_name,
        _designation_str(off.designation),
        bb.bank_name,
        bb.branch_name,
        str(bb.bank_code),
        off.account_number,
        int(off.num_days),
        off.telephone_number,
    )


def _write_centre_block(
    ws: object,
    start_row: int,
    school: School,
    exam_label: str,
    pairs: list[tuple[ExamCentreOfficial, School]],
    *,
    merge_title: bool,
) -> int:
    """Write title, headers, and rows; return next free row index."""
    ncols = len(_HEADER_LABELS)
    r = start_row
    title = f"Examination centre: {school.name} ({school.code}) · {exam_label}"
    if merge_title:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c1 = ws.cell(row=r, column=1, value=title)
    c1.font = Font(bold=True, size=13)
    c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[r].height = 28
    r += 1
    r += 1
    for i, h in enumerate(_HEADER_LABELS, start=1):
        ws.cell(row=r, column=i, value=h)
    _style_header_row(ws, r, ncols)
    r += 1
    for off, sch in pairs:
        vals = _data_values(off, sch)
        for i, v in enumerate(vals, start=1):
            ws.cell(row=r, column=i, value=v)
        _style_data_row(ws, r, ncols)
        r += 1
    return r


def _workbook_for_centre(
    school: School,
    exam_label: str,
    pairs: list[tuple[ExamCentreOfficial, School]],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Officials"
    _write_centre_block(ws, 1, school, exam_label, pairs, merge_title=True)
    _set_col_widths(ws, [12, 28, 24, 18, 22, 26, 12, 14, 6, 12])
    return wb


async def _load_examination(session: DBSessionDep, exam_id: int) -> Examination:
    ex = await session.get(Examination, exam_id)
    if ex is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Examination not found")
    return ex


def _base_official_query(
    examination_id: int,
    center_id: UUID | None,
):
    stmt = (
        select(ExamCentreOfficial, School)
        .join(School, School.id == ExamCentreOfficial.center_id)
        .where(ExamCentreOfficial.examination_id == examination_id)
        .options(selectinload(ExamCentreOfficial.bank_branch))
    )
    if center_id is not None:
        stmt = stmt.where(ExamCentreOfficial.center_id == center_id)
    return stmt.order_by(School.code.asc(), ExamCentreOfficial.full_name.asc())


@router.get("", response_model=AdminExamCentreOfficialListResponse)
async def admin_list_exam_centre_officials(
    session: DBSessionDep,
    _admin: SuperAdminDep,
    examination_id: int = Query(..., description="Examination id"),
    center_id: UUID | None = Query(None, description="Filter by examination centre (host school) id"),
    skip: int = Query(0, ge=0),
    limit: int = Query(_DEFAULT_LIST, ge=1, le=_MAX_LIST),
) -> AdminExamCentreOfficialListResponse:
    ex = await _load_examination(session, examination_id)
    exam_label = _examination_label(ex)

    count_stmt = select(func.count()).select_from(ExamCentreOfficial).where(
        ExamCentreOfficial.examination_id == examination_id
    )
    if center_id is not None:
        count_stmt = count_stmt.where(ExamCentreOfficial.center_id == center_id)
    total = int(await session.scalar(count_stmt) or 0)

    stmt = _base_official_query(examination_id, center_id).offset(skip).limit(limit)
    result = await session.execute(stmt)
    rows = result.all()

    items: list[AdminExamCentreOfficialRow] = []
    for off, school in rows:
        bb = off.bank_branch
        des = _designation_str(off.designation)
        items.append(
            AdminExamCentreOfficialRow(
                id=off.id,
                examination_id=examination_id,
                examination_label=exam_label,
                center_id=school.id,
                center_code=cast(str, school.code),
                center_name=cast(str, school.name),
                full_name=cast(str, off.full_name),
                designation=des,
                bank_branch_id=off.bank_branch_id,
                bank_code=cast(str, bb.bank_code),
                bank_name=cast(str, bb.bank_name),
                branch_name=cast(str, bb.branch_name),
                account_number=cast(str, off.account_number),
                num_days=int(off.num_days),
                telephone_number=cast(str, off.telephone_number),
                created_at=cast(datetime, off.created_at),
                updated_at=cast(datetime, off.updated_at),
            )
        )
    return AdminExamCentreOfficialListResponse(items=items, total=total)


@router.get("/export")
async def admin_export_exam_centre_officials(
    session: DBSessionDep,
    _admin: SuperAdminDep,
    examination_id: int = Query(..., description="Examination id"),
    layout: Literal["zip", "combined"] = Query("zip", description="zip = one workbook per centre in a zip; combined = one workbook"),
    center_id: UUID | None = Query(None, description="Optional: only this centre"),
) -> Response:
    ex = await _load_examination(session, examination_id)
    exam_label = _examination_label(ex)

    stmt = _base_official_query(examination_id, center_id)
    result = await session.execute(stmt)
    pairs: list[tuple[ExamCentreOfficial, School]] = list(result.all())
    if not pairs:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No exam officials found for this examination (and filter, if any).",
        )

    groups: dict[UUID, list[tuple[ExamCentreOfficial, School]]] = defaultdict(list)
    for off, school in pairs:
        groups[school.id].append((off, school))

    ordered = sorted(groups.items(), key=lambda kv: kv[1][0][1].code)

    exam_part = _safe_filename_part(f"exam_{examination_id}_{exam_label}")

    if layout == "zip":
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for _cid, plist in ordered:
                school = plist[0][1]
                wb = _workbook_for_centre(school, exam_label, plist)
                xbuf = io.BytesIO()
                wb.save(xbuf)
                fname = f"{_safe_filename_part(school.code)}-{_safe_filename_part(school.name)}.xlsx"
                zf.writestr(fname, xbuf.getvalue())
        payload = buf.getvalue()
        filename = f"{exam_part}_officials_by_centre.zip"
        media = "application/zip"
    else:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "All centres"
        r = 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(_HEADER_LABELS))
        tcell = ws.cell(
            row=r,
            column=1,
            value=f"Examination officials · {_examination_label(ex)} (all centres)",
        )
        tcell.font = Font(bold=True, size=14)
        tcell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 30
        r += 2
        for _cid, plist in ordered:
            school = plist[0][1]
            r = _write_centre_block(ws, r, school, exam_label, plist, merge_title=True)
            r += 2
        _set_col_widths(ws, [12, 28, 24, 18, 22, 26, 12, 14, 6, 12])
        xbuf = io.BytesIO()
        wb.save(xbuf)
        payload = xbuf.getvalue()
        filename = f"{exam_part}_officials_all_centres.xlsx"
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return Response(
        content=payload,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
