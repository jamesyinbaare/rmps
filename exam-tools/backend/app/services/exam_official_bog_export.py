"""Bank of Ghana (BoG) payment Excel export for exam centre officials."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from decimal import Decimal
from typing import cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import (
    ExamCentreOfficial,
    ExaminationCentre,
    ExamOfficialDesignation,
    ExaminationDesignationRate,
)
from app.schemas.timetable import TimetableDownloadFilter
from app.services.exam_official_compensation import compensation_for_official
from app.services.exam_official_designation import sort_officials_by_designation_then_name
from app.services.exam_official_export import designation_str, safe_filename_part
from app.services.finance_school_summary import subject_filter_filename_suffix

BOG_HEADER_LABELS = [
    "Serial",
    "Sort code",
    "Account number",
    "Name",
    "Designation",
    "Amount (GHS)",
]

BOG_COLUMN_WIDTHS = [10, 12, 22, 32, 24, 16]

SERIAL_COLUMN = 1
SORT_CODE_COLUMN = 2
ACCOUNT_COLUMN = 3
NAME_COLUMN = 4
DESIGNATION_COLUMN = 5
AMOUNT_COLUMN = 6

EXCEL_TEXT_FORMAT = "@"
AMOUNT_NUMBER_FORMAT = "#,##0.00"
GRAND_TOTAL_LABEL = "GRAND TOTAL"

_FILL_TITLE = PatternFill(fill_type="solid", fgColor="1E3A5F")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor="2E5077")
_FILL_ZEBRA_BASE = PatternFill(fill_type="solid", fgColor="FFFFFF")
_FILL_ZEBRA_ALT = PatternFill(fill_type="solid", fgColor="F8FAFC")
_FILL_TOTAL = PatternFill(fill_type="solid", fgColor="E8EEF4")

_FONT_TITLE = Font(bold=True, size=13, color="FFFFFF")
_FONT_HEADER = Font(bold=True, size=11, color="FFFFFF")
_FONT_DATA = Font(size=11, color="1F2937")
_FONT_NAME = Font(size=11, color="1F2937", bold=False)
_FONT_TOTAL = Font(bold=True, size=11, color="1F2937")

_SIDE_THIN = Side(style="thin", color="D1D5DB")
_SIDE_HEADER_BOTTOM = Side(style="medium", color="2E5077")
_SIDE_TOTAL_TOP = Side(style="medium", color="2E5077")


@dataclass(frozen=True)
class BogExportRow:
    serial: str
    sort_code: str
    account_number: str
    full_name: str
    designation: str
    amount: Decimal
    phone_number: str = ""
    incomplete_bank: bool = False
    reference_code: str = ""
    work_units: int | None = None
    num_days: int | None = None
    allocated_scripts: int | None = None
    # Counts aligned to script_paper_numbers passed into bog_workbook_bytes.
    paper_script_counts: tuple[int, ...] = ()


_FILL_INCOMPLETE = PatternFill(fill_type="solid", fgColor="FDE68A")

# Examiner BoG: Phone + Reference code after Name.
BOG_HEADER_LABELS_WITH_PHONE = [
    "Serial",
    "Sort code",
    "Account number",
    "Name",
    "Reference code",
    "Phone",
    "Designation",
    "Amount (GHS)",
]

BOG_COLUMN_WIDTHS_WITH_PHONE = [10, 12, 22, 28, 16, 14, 24, 16]

# Workforce BoG: Phone + scripts/entries + days after Name.
BOG_HEADER_LABELS_WORKFORCE = [
    "Serial",
    "Sort code",
    "Account number",
    "Name",
    "Phone",
    "Scripts",
    "Days",
    "Designation",
    "Amount (GHS)",
]

BOG_COLUMN_WIDTHS_WORKFORCE = [10, 12, 22, 28, 14, 10, 8, 22, 16]


def _bog_display_name(raw: str) -> str:
    return raw.strip().upper()


def _is_payable_official(
    off: ExamCentreOfficial,
    rates_by_designation: dict[ExamOfficialDesignation, ExaminationDesignationRate],
) -> Decimal | None:
    account = cast(str, off.account_number or "").strip()
    if not account:
        return None
    bb = off.bank_branch
    if bb is None:
        return None
    sort_code = str(bb.bank_code or "").strip()
    if not sort_code:
        return None
    comp = compensation_for_official(off, rates_by_designation)
    total = comp.total_payable_ghs
    if total is None or total <= 0:
        return None
    return total


def bog_export_rows(
    pairs: list[tuple[ExamCentreOfficial, ExaminationCentre]],
    rates_by_designation: dict[ExamOfficialDesignation, ExaminationDesignationRate],
) -> list[BogExportRow]:
    """Filter, sort, and assign 6-digit serial numbers for BoG export."""
    officials = [off for off, _ in pairs]
    sorted_officials = sort_officials_by_designation_then_name(officials)

    rows: list[BogExportRow] = []
    serial = 0
    for off in sorted_officials:
        amount = _is_payable_official(off, rates_by_designation)
        if amount is None:
            continue
        bb = off.bank_branch
        serial += 1
        rows.append(
            BogExportRow(
                serial=f"{serial:06d}",
                sort_code=str(bb.bank_code).strip(),
                account_number=cast(str, off.account_number).strip(),
                full_name=_bog_display_name(cast(str, off.full_name)),
                designation=designation_str(off.designation),
                amount=amount,
            )
        )
    return rows


def bog_grand_total(rows: list[BogExportRow]) -> Decimal:
    return sum((row.amount for row in rows), Decimal("0"))


def _filename_part(s: str) -> str:
    t = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", s.strip())
    return (t or "unknown")[:80]


def centre_bog_export_filename(
    center_code: str,
    center_name: str,
    subject_filter: TimetableDownloadFilter,
) -> str:
    suffix = subject_filter_filename_suffix(subject_filter)
    return f"{_filename_part(center_code)} {_filename_part(center_name)} BoG {suffix}.xlsx"


def exam_bog_export_filename(exam_part: str, slug: str) -> str:
    return f"{safe_filename_part(exam_part)}_bog_{safe_filename_part(slug)}.xlsx"


def _grid_border(*, header_bottom: bool = False, total_top: bool = False) -> Border:
    top = _SIDE_TOTAL_TOP if total_top else _SIDE_THIN
    bottom = _SIDE_HEADER_BOTTOM if header_bottom else _SIDE_THIN
    return Border(left=_SIDE_THIN, right=_SIDE_THIN, top=top, bottom=bottom)


def _header_alignment(column: int) -> Alignment:
    if column in (SERIAL_COLUMN, SORT_CODE_COLUMN, AMOUNT_COLUMN):
        return Alignment(horizontal="center", vertical="center", wrap_text=False)
    if column == ACCOUNT_COLUMN:
        return Alignment(horizontal="left", vertical="center", wrap_text=False)
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def _data_alignment(column: int, *, amount_column: int = AMOUNT_COLUMN) -> Alignment:
    if column == SERIAL_COLUMN:
        return Alignment(horizontal="center", vertical="center", wrap_text=False)
    if column in (SORT_CODE_COLUMN, ACCOUNT_COLUMN):
        return Alignment(horizontal="left", vertical="center", wrap_text=False)
    if column == amount_column:
        return Alignment(horizontal="right", vertical="center", wrap_text=False)
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _style_title_row(ws: object, row: int, ncols: int, title: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = _FILL_TITLE
    cell.font = _FONT_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[row].height = 32


def _style_header_row(ws: object, row: int, ncols: int, *, amount_column: int = AMOUNT_COLUMN) -> None:
    border = _grid_border(header_bottom=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.border = border
        if col == amount_column:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        else:
            cell.alignment = _header_alignment(col)
    ws.row_dimensions[row].height = 22


def _write_text_cell(ws: object, row: int, column: int, value: str) -> None:
    cell = ws.cell(row=row, column=column, value=value)
    cell.number_format = EXCEL_TEXT_FORMAT


def _write_amount_cell(ws: object, row: int, column: int, value: Decimal) -> None:
    cell = ws.cell(row=row, column=column, value=float(value))
    cell.number_format = AMOUNT_NUMBER_FORMAT


def _write_int_cell(ws: object, row: int, column: int, value: int) -> None:
    cell = ws.cell(row=row, column=column, value=value)
    cell.number_format = "0"


def _style_data_row(
    ws: object,
    row: int,
    ncols: int,
    *,
    stripe: bool,
    incomplete: bool = False,
    name_column: int = NAME_COLUMN,
    amount_column: int = AMOUNT_COLUMN,
) -> None:
    if incomplete:
        fill = _FILL_INCOMPLETE
    else:
        fill = _FILL_ZEBRA_ALT if stripe else _FILL_ZEBRA_BASE
    border = _grid_border()
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = _FONT_NAME if col == name_column else _FONT_DATA
        cell.border = border
        cell.alignment = _data_alignment(col, amount_column=amount_column)
    ws.row_dimensions[row].height = 20


def _apply_sheet_finish(ws: object, header_row: int, ncols: int, last_data_row: int) -> None:
    ws.freeze_panes = f"A{header_row + 1}"
    if last_data_row >= header_row:
        last_col = get_column_letter(ncols)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_data_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"${header_row}:${header_row}"


def bog_workbook_bytes(
    pairs: list[tuple[ExamCentreOfficial, ExaminationCentre]],
    rates_by_designation: dict[ExamOfficialDesignation, ExaminationDesignationRate],
    *,
    title: str | None = None,
    prebuilt_rows: list[BogExportRow] | None = None,
    include_phone: bool = False,
    include_workforce_detail: bool = False,
    work_unit_label: str = "Scripts",
    script_paper_numbers: list[int] | None = None,
) -> bytes:
    rows = prebuilt_rows if prebuilt_rows is not None else bog_export_rows(pairs, rates_by_designation)
    total = bog_grand_total(rows)
    papers = list(script_paper_numbers or [])

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "BoG payment"
    if include_workforce_detail:
        headers = list(BOG_HEADER_LABELS_WORKFORCE)
        headers[5] = work_unit_label
        widths = list(BOG_COLUMN_WIDTHS_WORKFORCE)
        name_col = 4
        phone_col = 5
        work_units_col = 6
        days_col = 7
        designation_col = 8
        amount_col = 9
        reference_col = None
        allocated_col = None
        paper_cols: list[int] = []
        if script_paper_numbers is not None:
            work_unit_idx = headers.index(work_unit_label)
            del headers[work_unit_idx]
            del widths[work_unit_idx]
            for i, paper in enumerate(papers):
                headers.insert(work_unit_idx + i, f"Paper {paper} scripts")
                widths.insert(work_unit_idx + i, 12)
            work_units_col = None
            paper_cols = [work_unit_idx + 1 + i for i in range(len(papers))]
            days_col = headers.index("Days") + 1
            designation_col = headers.index("Designation") + 1
            amount_col = headers.index("Amount (GHS)") + 1
    elif include_phone:
        headers = list(BOG_HEADER_LABELS_WITH_PHONE)
        widths = list(BOG_COLUMN_WIDTHS_WITH_PHONE)
        name_col = 4
        reference_col = 5
        phone_col = 6
        work_units_col = None
        days_col = None
        designation_col = 7
        amount_col = 8
        allocated_col = None
        paper_cols: list[int] = []
        # None = omit (e.g. T&T BoG); list (even empty) = Allocated scripts + Paper N columns.
        if script_paper_numbers is not None:
            insert_at = headers.index("Amount (GHS)")
            headers.insert(insert_at, "Allocated scripts")
            widths.insert(insert_at, 12)
            allocated_col = insert_at + 1
            for i, paper in enumerate(papers):
                headers.insert(insert_at + 1 + i, f"Paper {paper} scripts")
                widths.insert(insert_at + 1 + i, 12)
            paper_cols = [allocated_col + 1 + i for i in range(len(papers))]
            amount_col = headers.index("Amount (GHS)") + 1
            designation_col = headers.index("Designation") + 1
    else:
        headers = list(BOG_HEADER_LABELS)
        widths = list(BOG_COLUMN_WIDTHS)
        name_col = NAME_COLUMN
        reference_col = None
        phone_col = None
        work_units_col = None
        days_col = None
        designation_col = DESIGNATION_COLUMN
        amount_col = AMOUNT_COLUMN
        allocated_col = None
        paper_cols = []
    ncols = len(headers)
    start_row = 1

    if title:
        _style_title_row(ws, start_row, ncols, title)
        start_row += 1

    header_row = start_row
    for col, label in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=label)
    _style_header_row(ws, header_row, ncols, amount_column=amount_col)

    data_row = header_row + 1
    for idx, row in enumerate(rows):
        r = data_row + idx
        _write_text_cell(ws, r, SERIAL_COLUMN, row.serial)
        _write_text_cell(ws, r, SORT_CODE_COLUMN, row.sort_code)
        _write_text_cell(ws, r, ACCOUNT_COLUMN, row.account_number)
        _write_text_cell(ws, r, name_col, row.full_name)
        if reference_col is not None:
            _write_text_cell(ws, r, reference_col, row.reference_code or "")
        if phone_col is not None:
            _write_text_cell(ws, r, phone_col, row.phone_number or "")
        if work_units_col is not None:
            _write_int_cell(ws, r, work_units_col, int(row.work_units or 0))
        if days_col is not None:
            _write_int_cell(ws, r, days_col, int(row.num_days or 0))
        _write_text_cell(ws, r, designation_col, row.designation)
        if allocated_col is not None:
            _write_int_cell(ws, r, allocated_col, int(row.allocated_scripts or 0))
        for col_i, paper_col in enumerate(paper_cols):
            count = row.paper_script_counts[col_i] if col_i < len(row.paper_script_counts) else 0
            _write_int_cell(ws, r, paper_col, int(count))
        _write_amount_cell(ws, r, amount_col, row.amount)
        _style_data_row(
            ws,
            r,
            ncols,
            stripe=idx % 2 == 1,
            incomplete=row.incomplete_bank,
            name_column=name_col,
            amount_column=amount_col,
        )

    total_row = data_row + len(rows)
    total_border = _grid_border(total_top=True)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = _FILL_TOTAL
        cell.font = _FONT_TOTAL
        cell.border = total_border
    ws.cell(row=total_row, column=designation_col, value=GRAND_TOTAL_LABEL)
    _write_amount_cell(ws, total_row, amount_col, total)
    ws.cell(row=total_row, column=designation_col).alignment = Alignment(
        horizontal="right", vertical="center"
    )
    ws.cell(row=total_row, column=amount_col).font = _FONT_TOTAL
    ws.cell(row=total_row, column=amount_col).alignment = Alignment(
        horizontal="right", vertical="center"
    )
    ws.row_dimensions[total_row].height = 22

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_data_row = total_row - 1 if rows else header_row
    _apply_sheet_finish(ws, header_row, ncols, last_data_row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
