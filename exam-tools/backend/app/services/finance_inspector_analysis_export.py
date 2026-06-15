"""Excel export for per-centre external inspector analysis."""

from __future__ import annotations

import io
import re
from datetime import UTC, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.examination import FinanceCentreInspectorAnalysisRow
from app.schemas.timetable import TimetableDownloadFilter
from app.services.finance_school_summary import subject_filter_filename_suffix

SHEET_NAME = "Inspector analysis"
LEGEND_SHEET_NAME = "Legend"

FULL_HEADER_LABELS = [
    "Centre code",
    "Centre name",
    "Candidates",
    "Exam days",
    "Max assigned days",
    "Days variance",
    "Required",
    "Paid inspectors",
    "Posted inspectors",
    "Unique total",
    "In both",
    "Posted not on payroll",
    "Staffing variance",
    "Candidates/inspector",
    "Pay at exam days (GHS)",
    "Pay at assigned (GHS)",
    "Days pay variance (GHS)",
    "Pay at posted (GHS)",
    "Payroll vs posted (GHS)",
    "Total pay (GHS)",
]

STAFFING_HEADER_LABELS = [
    "Centre code",
    "Centre name",
    "Candidates",
    "Exam days",
    "Required",
    "Paid inspectors",
    "Posted inspectors",
    "Unique total",
    "In both",
    "Posted not on payroll",
    "Staffing variance",
    "Candidates/inspector",
    "Total pay (GHS)",
]

PAY_VARIANCE_HEADER_LABELS = [
    "Centre code",
    "Centre name",
    "Exam days",
    "Max assigned days",
    "Days variance",
    "Paid inspectors",
    "Posted inspectors",
    "Roster pay (GHS)",
    "Pay at exam days (GHS)",
    "Pay at assigned (GHS)",
    "Days pay variance (GHS)",
    "Pay at posted (GHS)",
    "Payroll vs posted (GHS)",
]

STANDARD_FULL_HEADER_LABELS = [h for h in FULL_HEADER_LABELS if h != "Posted not on payroll"]
STANDARD_STAFFING_HEADER_LABELS = [h for h in STAFFING_HEADER_LABELS if h != "Posted not on payroll"]

HEADER_LABELS = STANDARD_FULL_HEADER_LABELS

_FILL_TITLE = PatternFill(fill_type="solid", fgColor="1E3A5F")
_FILL_PREAMBLE = PatternFill(fill_type="solid", fgColor="F4F6F8")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor="2E5077")
_FILL_ZEBRA_BASE = PatternFill(fill_type="solid", fgColor="FFFFFF")
_FILL_ZEBRA_ALT = PatternFill(fill_type="solid", fgColor="F8FAFC")
_FILL_TOTALS = PatternFill(fill_type="solid", fgColor="EEF2FF")
_FILL_VARIANCE_OVER = PatternFill(fill_type="solid", fgColor="FEE2E2")
_FILL_VARIANCE_SHORT = PatternFill(fill_type="solid", fgColor="FEF3C7")
_FILL_VARIANCE_MATCH = PatternFill(fill_type="solid", fgColor="ECFDF5")
_FILL_PAYROLL_GAP = PatternFill(fill_type="solid", fgColor="FFFBEB")
_FILL_KPI = PatternFill(fill_type="solid", fgColor="F0F9FF")
_FILL_GROUP_CENTRE = PatternFill(fill_type="solid", fgColor="475569")
_FILL_GROUP_SCALE = PatternFill(fill_type="solid", fgColor="1D4ED8")
_FILL_GROUP_ROSTER = PatternFill(fill_type="solid", fgColor="6D28D9")
_FILL_GROUP_STAFFING = PatternFill(fill_type="solid", fgColor="047857")
_FILL_GROUP_PAY = PatternFill(fill_type="solid", fgColor="B45309")
_FILL_LEGEND_HEADER = PatternFill(fill_type="solid", fgColor="334155")

_FONT_TITLE = Font(bold=True, size=14, color="FFFFFF")
_FONT_PREAMBLE = Font(size=11, color="374151")
_FONT_HEADER = Font(bold=True, size=11, color="FFFFFF")
_FONT_DATA = Font(size=11, color="1F2937")
_FONT_TOTALS = Font(bold=True, size=11, color="1F2937")
_FONT_KPI_LABEL = Font(bold=True, size=10, color="475569")
_FONT_KPI_VALUE = Font(bold=True, size=12, color="0F172A")
_FONT_GROUP = Font(bold=True, size=10, color="FFFFFF")
_FONT_VARIANCE_OVER = Font(bold=True, size=11, color="B91C1C")
_FONT_VARIANCE_SHORT = Font(bold=True, size=11, color="B45309")
_FONT_VARIANCE_MATCH = Font(bold=True, size=11, color="047857")
_FONT_PAYROLL_GAP = Font(bold=True, size=11, color="B45309")

_SIDE_THIN = Side(style="thin", color="D1D5DB")
_SIDE_HEADER_BOTTOM = Side(style="medium", color="2E5077")
_SIDE_MEDIUM = Side(style="medium", color="94A3B8")


def inspector_analysis_export_filename(
    exam_label: str,
    subject_filter: TimetableDownloadFilter,
    *,
    export_variant: str = "full",
    export_style: str = "standard",
) -> str:
    def part(s: str) -> str:
        t = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", s.strip())
        return (t or "unknown")[:80]

    suffix = subject_filter_filename_suffix(subject_filter)
    variant_suffix = {
        "staffing": "staffing",
        "pay_variance": "pay-variance",
    }.get(export_variant, "full")
    style_suffix = " formatted" if export_style == "rich" else ""
    return f"{part(exam_label)} inspector-analysis {suffix} {variant_suffix}{style_suffix}.xlsx"


def _grid_border(*, header_bottom: bool = False) -> Border:
    bottom = _SIDE_HEADER_BOTTOM if header_bottom else _SIDE_THIN
    return Border(left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=bottom)


def _posted_not_on_payroll(row: FinanceCentreInspectorAnalysisRow) -> int:
    return max(0, row.posted_inspector_count - row.inspectors_in_both)


def _variance_fill(variance: int | float) -> PatternFill:
    if variance > 0:
        return _FILL_VARIANCE_OVER
    if variance < 0:
        return _FILL_VARIANCE_SHORT
    return _FILL_VARIANCE_MATCH


def _variance_font(variance: int | float) -> Font:
    if variance > 0:
        return _FONT_VARIANCE_OVER
    if variance < 0:
        return _FONT_VARIANCE_SHORT
    return _FONT_VARIANCE_MATCH


def _headers_for_variant(export_variant: str, *, export_style: str) -> list[str]:
    rich = export_style == "rich"
    if export_variant == "staffing":
        return STAFFING_HEADER_LABELS if rich else STANDARD_STAFFING_HEADER_LABELS
    if export_variant == "pay_variance":
        return PAY_VARIANCE_HEADER_LABELS
    return FULL_HEADER_LABELS if rich else STANDARD_FULL_HEADER_LABELS


def _group_spans(headers: list[str], export_variant: str) -> list[tuple[str, int, int, PatternFill]]:
    """Return (label, start_col, end_col, fill) 1-based inclusive."""
    label_to_col = {label: idx + 1 for idx, label in enumerate(headers)}

    def span(label: str, start: str, end: str, fill: PatternFill) -> tuple[str, int, int, PatternFill]:
        return label, label_to_col[start], label_to_col[end], fill

    if export_variant == "staffing":
        return [
            span("Centre", "Centre code", "Centre name", _FILL_GROUP_CENTRE),
            span("Scale", "Candidates", "Required", _FILL_GROUP_SCALE),
            span("Roster", "Paid inspectors", "Posted not on payroll", _FILL_GROUP_ROSTER)
            if "Posted not on payroll" in label_to_col
            else span("Roster", "Paid inspectors", "In both", _FILL_GROUP_ROSTER),
            span("Staffing", "Staffing variance", "Candidates/inspector", _FILL_GROUP_STAFFING),
            span("Pay", "Total pay (GHS)", "Total pay (GHS)", _FILL_GROUP_PAY),
        ]
    if export_variant == "pay_variance":
        return [
            span("Centre", "Centre code", "Centre name", _FILL_GROUP_CENTRE),
            span("Scale", "Exam days", "Days variance", _FILL_GROUP_SCALE),
            span("Roster", "Paid inspectors", "Posted inspectors", _FILL_GROUP_ROSTER),
            span("Pay", "Roster pay (GHS)", "Payroll vs posted (GHS)", _FILL_GROUP_PAY),
        ]
    return [
        span("Centre", "Centre code", "Centre name", _FILL_GROUP_CENTRE),
        span("Scale", "Candidates", "Days variance", _FILL_GROUP_SCALE),
        span("Roster", "Required", "Posted not on payroll", _FILL_GROUP_ROSTER)
        if "Posted not on payroll" in label_to_col
        else span("Roster", "Required", "In both", _FILL_GROUP_ROSTER),
        span("Staffing", "Staffing variance", "Candidates/inspector", _FILL_GROUP_STAFFING),
        span("Pay", "Pay at exam days (GHS)", "Total pay (GHS)", _FILL_GROUP_PAY),
    ]


def _row_values(
    row: FinanceCentreInspectorAnalysisRow,
    export_variant: str = "full",
    *,
    export_style: str = "standard",
) -> list[object]:
    rich = export_style == "rich"
    payroll_gap = _posted_not_on_payroll(row)
    if export_variant == "staffing":
        values: list[object] = [
            row.center_code,
            row.center_name,
            row.total_candidates,
            row.exam_days,
            row.inspectors_required,
            row.external_inspector_count,
            row.posted_inspector_count,
            row.unique_inspector_count,
            row.inspectors_in_both,
        ]
        if rich:
            values.append(payroll_gap)
        values.extend(
            [
                row.paid_inspector_variance,
                row.candidates_per_paid_inspector if row.candidates_per_paid_inspector is not None else "",
                float(row.total_inspector_pay_ghs),
            ]
        )
        return values
    if export_variant == "pay_variance":
        return [
            row.center_code,
            row.center_name,
            row.exam_days,
            row.max_inspector_assigned_days,
            row.assigned_days_variance,
            row.external_inspector_count,
            row.posted_inspector_count,
            float(row.total_inspector_pay_ghs),
            float(row.pay_at_exam_days_ghs),
            float(row.pay_at_assigned_days_ghs),
            float(row.days_pay_variance_ghs),
            float(row.pay_at_posted_count_ghs),
            float(row.payroll_vs_posted_variance_ghs),
        ]
    values = [
        row.center_code,
        row.center_name,
        row.total_candidates,
        row.exam_days,
        row.max_inspector_assigned_days,
        row.assigned_days_variance,
        row.inspectors_required,
        row.external_inspector_count,
        row.posted_inspector_count,
        row.unique_inspector_count,
        row.inspectors_in_both,
    ]
    if rich:
        values.append(payroll_gap)
    values.extend(
        [
            row.paid_inspector_variance,
            row.candidates_per_paid_inspector if row.candidates_per_paid_inspector is not None else "",
            float(row.pay_at_exam_days_ghs),
            float(row.pay_at_assigned_days_ghs),
            float(row.days_pay_variance_ghs),
            float(row.pay_at_posted_count_ghs),
            float(row.payroll_vs_posted_variance_ghs),
            float(row.total_inspector_pay_ghs),
        ]
    )
    return values


def _variance_col_indices(headers: list[str]) -> dict[str, int]:
    out: dict[str, int] = {}
    for key, label in (
        ("days", "Days variance"),
        ("staffing", "Staffing variance"),
        ("days_pay", "Days pay variance (GHS)"),
        ("payroll_vs_posted", "Payroll vs posted (GHS)"),
    ):
        if label in headers:
            out[key] = headers.index(label) + 1
    return out


def _money_col_indices(headers: list[str]) -> set[int]:
    money_labels = {
        "Pay at exam days (GHS)",
        "Pay at assigned (GHS)",
        "Days pay variance (GHS)",
        "Pay at posted (GHS)",
        "Payroll vs posted (GHS)",
        "Total pay (GHS)",
        "Roster pay (GHS)",
    }
    return {headers.index(label) + 1 for label in headers if label in money_labels}


def _numeric_col_indices(headers: list[str], money_cols: set[int]) -> set[int]:
    text_cols = {"Centre code", "Centre name"}
    return {i + 1 for i, h in enumerate(headers) if h not in text_cols} | money_cols


def _payroll_gap_col(headers: list[str]) -> int | None:
    if "Posted not on payroll" not in headers:
        return None
    return headers.index("Posted not on payroll") + 1


def _style_merged_row(
    ws: Worksheet,
    row: int,
    ncols: int,
    *,
    value: str,
    fill: PatternFill,
    font: Font,
    height: int = 24,
) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=value)
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = height


def _write_kpi_strip(
    ws: Worksheet,
    row: int,
    ncols: int,
    rows: list[FinanceCentreInspectorAnalysisRow],
    totals: FinanceCentreInspectorAnalysisRow,
    export_variant: str,
    candidates_per_inspector: int,
) -> None:
    over_staffed = sum(1 for r in rows if r.paid_inspector_variance > 0)
    under_staffed = sum(1 for r in rows if r.paid_inspector_variance < 0)
    payroll_gaps = sum(1 for r in rows if _posted_not_on_payroll(r) > 0)
    days_pay_over = sum(1 for r in rows if float(r.days_pay_variance_ghs) > 0)

    if export_variant == "staffing":
        kpis = [
            ("Rule", f"1 / {candidates_per_inspector} candidates"),
            ("Paid / required", f"{totals.external_inspector_count} / {totals.inspectors_required}"),
            ("Over-staffed centres", str(over_staffed)),
            ("Under-staffed centres", str(under_staffed)),
            ("Payroll gap centres", str(payroll_gaps)),
            ("Total pay (GHS)", f"{float(totals.total_inspector_pay_ghs):,.2f}"),
        ]
    elif export_variant == "pay_variance":
        kpis = [
            ("Exam / max assigned days", f"{totals.exam_days} / {totals.max_inspector_assigned_days}"),
            ("Days variance (total)", str(totals.assigned_days_variance)),
            ("Days pay over centres", str(days_pay_over)),
            ("Days pay var. (GHS)", f"{float(totals.days_pay_variance_ghs):+,.2f}"),
            ("Payroll vs posted (GHS)", f"{float(totals.payroll_vs_posted_variance_ghs):+,.2f}"),
            ("Roster pay (GHS)", f"{float(totals.total_inspector_pay_ghs):,.2f}"),
        ]
    else:
        kpis = [
            ("Candidates", f"{totals.total_candidates:,}"),
            ("Paid / required", f"{totals.external_inspector_count} / {totals.inspectors_required}"),
            ("Over / under centres", f"{over_staffed} / {under_staffed}"),
            ("Days pay var. (GHS)", f"{float(totals.days_pay_variance_ghs):+,.2f}"),
            ("Payroll gap centres", str(payroll_gaps)),
            ("Total pay (GHS)", f"{float(totals.total_inspector_pay_ghs):,.2f}"),
        ]

    span = max(1, ncols // len(kpis))
    col = 1
    for label, value in kpis:
        end_col = min(col + span - 1, ncols)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
        cell = ws.cell(row=row, column=col, value=f"{label}: {value}")
        cell.fill = _FILL_KPI
        cell.font = _FONT_KPI_VALUE
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(
            left=_SIDE_THIN,
            right=_SIDE_THIN,
            top=_SIDE_MEDIUM,
            bottom=_SIDE_MEDIUM,
        )
        col = end_col + 1
        if col > ncols:
            break
    ws.row_dimensions[row].height = 28


def _write_group_header_row(ws: Worksheet, row: int, headers: list[str], export_variant: str) -> None:
    ncols = len(headers)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col, value="")
        cell.fill = _FILL_PREAMBLE
        cell.border = _grid_border(header_bottom=True)
    for label, start_col, end_col, fill in _group_spans(headers, export_variant):
        if start_col > end_col:
            continue
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        cell = ws.cell(row=row, column=start_col, value=label)
        cell.fill = fill
        cell.font = _FONT_GROUP
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _grid_border(header_bottom=True)
    ws.row_dimensions[row].height = 22


def _write_legend_sheet(wb: Workbook, export_variant: str, candidates_per_inspector: int) -> None:
    ws = wb.create_sheet(LEGEND_SHEET_NAME[:31])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 64

    rows: list[tuple[str, str]] = [
        ("Report guide", "How to read this workbook"),
        ("", ""),
        ("Colour — red fill", "Higher than baseline (over-staffed, extra days, or higher pay)."),
        ("Colour — amber fill", "Lower than baseline (under-staffed, fewer days, or lower pay)."),
        ("Colour — green fill", "Exact match with baseline."),
        ("Colour — amber row tint", "Centre has posted inspectors not on the payroll roster."),
        ("", ""),
        ("Required headcount", f"ceil(candidates ÷ {candidates_per_inspector})"),
        ("Staffing variance", "Paid unique phones minus required. Positive = over-staffed."),
        ("Posted not on payroll", "Posted system inspectors whose phones are not on payroll."),
        ("Days variance", "Max assigned roster days minus timetable exam days."),
        ("Days pay variance (GHS)", "Pay at assigned roster days minus pay at exam days."),
        ("Payroll vs posted (GHS)", "Actual roster pay minus hypothetical pay for posted headcount."),
    ]
    if export_variant == "staffing":
        rows = [r for r in rows if "Days " not in r[0] or r[0] in {"Days variance"}]
    elif export_variant == "pay_variance":
        rows = [r for r in rows if r[0] not in {"Required headcount", "Staffing variance", "Posted not on payroll"}]

    for idx, (label, detail) in enumerate(rows, start=1):
        a = ws.cell(row=idx, column=1, value=label)
        b = ws.cell(row=idx, column=2, value=detail)
        if idx == 1:
            a.fill = _FILL_LEGEND_HEADER
            b.fill = _FILL_LEGEND_HEADER
            a.font = Font(bold=True, size=12, color="FFFFFF")
            b.font = Font(bold=True, size=12, color="FFFFFF")
        elif label.startswith("Colour"):
            swatch = _FILL_VARIANCE_OVER
            if "amber" in label:
                swatch = _FILL_VARIANCE_SHORT if "fill" in label else _FILL_PAYROLL_GAP
            elif "green" in label:
                swatch = _FILL_VARIANCE_MATCH
            a.fill = swatch
            a.font = _FONT_DATA
            b.font = _FONT_DATA
        elif label:
            a.font = _FONT_KPI_LABEL
            b.font = _FONT_DATA
            b.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[idx].height = 20 if not detail or len(detail) < 60 else 36


def _auto_column_widths(ws: Worksheet, headers: list[str], data_start: int, last_row: int) -> None:
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in range(data_start, last_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 40)


def _apply_print_setup(ws: Worksheet, header_row: int) -> None:
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"{header_row}:{header_row}"


def inspector_analysis_workbook_bytes(
    rows: list[FinanceCentreInspectorAnalysisRow],
    *,
    totals: FinanceCentreInspectorAnalysisRow,
    exam_label: str,
    subject_filter: TimetableDownloadFilter,
    candidates_per_inspector: int = 300,
    export_variant: str = "full",
    export_style: str = "standard",
) -> bytes:
    rich = export_style == "rich"
    headers = _headers_for_variant(export_variant, export_style=export_style)
    variance_cols = _variance_col_indices(headers)
    money_cols = _money_col_indices(headers)
    numeric_cols = _numeric_col_indices(headers, money_cols)
    payroll_gap_col = _payroll_gap_col(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME[:31]
    ncols = len(headers)

    title_suffix = {
        "staffing": "Staffing",
        "pay_variance": "Pay variance",
    }.get(export_variant, "Full report")
    style_label = " — formatted" if rich else ""

    title_row = 1
    _style_merged_row(
        ws,
        title_row,
        ncols,
        value=f"Inspector analysis — {exam_label} ({title_suffix}{style_label})",
        fill=_FILL_TITLE,
        font=_FONT_TITLE,
        height=34,
    )

    scope_label = subject_filter_filename_suffix(subject_filter)
    preamble_row = 2
    _style_merged_row(
        ws,
        preamble_row,
        ncols,
        value=f"Subject scope: {scope_label} · Ratio: {candidates_per_inspector} candidates per inspector",
        fill=_FILL_PREAMBLE,
        font=_FONT_PREAMBLE,
    )

    next_row = preamble_row + 1
    if rich:
        generated = datetime.now(tz=UTC).strftime("%Y-%m-%d %H:%M UTC")
        _style_merged_row(
            ws,
            next_row,
            ncols,
            value=f"Generated {generated} · Variance colours: red = over, amber = under, green = match",
            fill=_FILL_PREAMBLE,
            font=_FONT_PREAMBLE,
            height=20,
        )
        next_row += 1
        _write_kpi_strip(ws, next_row, ncols, rows, totals, export_variant, candidates_per_inspector)
        next_row += 1
        ws.row_dimensions[next_row].height = 6
        next_row += 1
        group_row = next_row
        _write_group_header_row(ws, group_row, headers, export_variant)
        header_row = group_row + 1
    else:
        ratio_row = next_row
        _style_merged_row(
            ws,
            ratio_row,
            ncols,
            value=f"Ratio: {candidates_per_inspector} candidates per inspector",
            fill=_FILL_PREAMBLE,
            font=_FONT_PREAMBLE,
        )
        header_row = ratio_row + 2

    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.border = _grid_border(header_bottom=True)
        cell.alignment = Alignment(
            horizontal="right" if col in numeric_cols else "left",
            vertical="center",
            wrap_text=True,
        )
    ws.row_dimensions[header_row].height = 40 if rich else 36

    data_start = header_row + 1
    all_rows = [*rows, totals]
    for idx, stat_row in enumerate(all_rows):
        excel_row = data_start + idx
        is_totals = stat_row.center_code == "TOTAL"
        stripe = idx % 2 == 1
        has_payroll_gap = not is_totals and _posted_not_on_payroll(stat_row) > 0
        values = _row_values(stat_row, export_variant, export_style=export_style)
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=value)
            variance_key = None
            if col == variance_cols.get("days"):
                variance_key = "days"
            elif col == variance_cols.get("staffing"):
                variance_key = "staffing"
            elif col == variance_cols.get("days_pay"):
                variance_key = "days_pay"
            elif col == variance_cols.get("payroll_vs_posted"):
                variance_key = "payroll_vs_posted"

            if variance_key:
                numeric_value = float(value) if variance_key in {"days_pay", "payroll_vs_posted"} else int(value)
                cell.fill = _variance_fill(numeric_value)
                cell.font = _FONT_TOTALS if is_totals else _variance_font(numeric_value)
                if variance_key in {"days_pay", "payroll_vs_posted"}:
                    cell.number_format = '+#,##0.00;-#,##0.00;"—"'
                else:
                    cell.number_format = '+0;-0;"—"'
            elif payroll_gap_col is not None and col == payroll_gap_col and int(value or 0) > 0:
                cell.fill = _FILL_PAYROLL_GAP
                cell.font = _FONT_PAYROLL_GAP if not is_totals else _FONT_TOTALS
            elif is_totals:
                cell.fill = _FILL_TOTALS
                cell.font = _FONT_TOTALS
            elif has_payroll_gap and rich:
                cell.fill = _FILL_PAYROLL_GAP if col <= 2 else (_FILL_ZEBRA_ALT if stripe else _FILL_ZEBRA_BASE)
                cell.font = _FONT_DATA
            else:
                cell.fill = _FILL_ZEBRA_ALT if stripe else _FILL_ZEBRA_BASE
                cell.font = _FONT_TOTALS if is_totals else _FONT_DATA

            cell.border = _grid_border()
            cell.alignment = Alignment(
                horizontal="right" if col in numeric_cols else "left",
                vertical="center",
            )
            if col in money_cols and variance_key not in {"days_pay", "payroll_vs_posted"}:
                cell.number_format = "#,##0.00"
            elif col in numeric_cols and variance_key is None:
                cell.number_format = "#,##0"

    ws.freeze_panes = ws.cell(row=data_start, column=1).coordinate
    last_row = data_start + len(all_rows) - 1
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{last_row}"

    if rich:
        _auto_column_widths(ws, headers, data_start, last_row)
        _apply_print_setup(ws, header_row)
        _write_legend_sheet(wb, export_variant, candidates_per_inspector)
    else:
        default_widths = [14, 32, 12, 10, 14, 12, 10, 14, 16, 12, 10, 14, 16, 16, 16, 16, 16, 18, 14]
        for col in range(1, ncols + 1):
            width = default_widths[col - 1] if col - 1 < len(default_widths) else 12
            ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
