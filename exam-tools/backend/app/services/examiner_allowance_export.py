"""Excel export helpers for examiner allowances."""

from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import Examiner, Examination, ExaminerType
from app.schemas.admin_examiner_allowance import AdminExaminerAllowanceRow
from app.services.exam_official_export import examination_label, safe_filename_part
from app.services.examiner_allowance_list import MarkingScriptSourceModes, examiners_to_admin_rows
from app.services.examiner_allocated_booklets import AllocatedBookletsMap
from app.services.examiner_compensation import (
    MarkingRateMap,
    RoleAllowanceMap,
    TravelRateMap,
    TravelRoleFactorMap,
    TravelZoneMap,
    TravelZoneNameMap,
)
from app.services.examiner_invitation import _examiner_type_label

# Core columns always included (order matters).
CORE_HEADER_LABELS = [
    "Code",
    "Name",
    "Role",
    "Region",
    "Subjects",
    "Bank",
    "Branch",
    "Bank code",
    "Account",
    "Bank status",
    "Phone",
    "Responsibility (GHS)",
    "Inconvenience (GHS)",
    "Chief Examiner's Report (GHS)",
    "Vetting of Scripts (GHS)",
    "Vetting tax (GHS)",
    "Vetting net (GHS)",
    "Internal Commuting (GHS)",
    "Marking (GHS)",
    "Marking tax (GHS)",
    "Marking net (GHS)",
    "Allocated scripts",
    "T & T (GHS)",
    "Payout T&T & commuting (GHS)",
    "Payout allowances & marking (GHS)",
    "Total payable (GHS)",
]

CORE_COLUMN_WIDTHS = [
    10,
    26,
    22,
    16,
    28,
    26,
    12,
    16,
    14,
    14,
    14,
    14,
    16,
    16,
    20,
    18,
    14,
    14,
    18,
    14,
    14,
    14,
    14,
    18,
    22,
    16,
]

# Optional extras officers may tick before download.
OPTIONAL_EXPORT_FIELDS: dict[str, tuple[str, int]] = {
    # key -> (header label, default column width)
    "subject_names": ("Subject names", 28),
    "travel_zone": ("Travel zone", 16),
}

OPTIONAL_EXPORT_FIELD_KEYS = frozenset(OPTIONAL_EXPORT_FIELDS)

# Back-compat alias used by older imports/tests.
HEADER_LABELS = list(CORE_HEADER_LABELS)
COLUMN_WIDTHS = list(CORE_COLUMN_WIDTHS)

EXCEL_TEXT_FORMAT = "@"
AMOUNT_NUMBER_FORMAT = "#,##0.00"

_FILL_TITLE = PatternFill(fill_type="solid", fgColor="1E3A5F")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor="2E5077")
_FILL_ZEBRA_BASE = PatternFill(fill_type="solid", fgColor="FFFFFF")
_FILL_ZEBRA_ALT = PatternFill(fill_type="solid", fgColor="F8FAFC")
_FILL_INCOMPLETE = PatternFill(fill_type="solid", fgColor="FDE68A")  # amber-200

_FONT_TITLE = Font(bold=True, size=13, color="FFFFFF")
_FONT_HEADER = Font(bold=True, size=11, color="FFFFFF")
_FONT_DATA = Font(size=11, color="1F2937")

_SIDE_THIN = Side(style="thin", color="D1D5DB")
_SIDE_HEADER_BOTTOM = Side(style="medium", color="2E5077")


def parse_include_fields(raw: str | None) -> frozenset[str]:
    """Parse comma-separated include_fields query; unknown keys ignored."""
    if raw is None or not str(raw).strip():
        return frozenset()
    selected: set[str] = set()
    for part in str(raw).split(","):
        key = part.strip()
        if key in OPTIONAL_EXPORT_FIELD_KEYS:
            selected.add(key)
    return frozenset(selected)


def _role_label(examiner_type: str) -> str:
    try:
        return _examiner_type_label(ExaminerType(examiner_type))
    except ValueError:
        return examiner_type


def examiner_export_filename(exam: Examination) -> str:
    return f"{safe_filename_part(examination_label(exam))}_examiner_allowances.xlsx"


def bank_account_incomplete(item: AdminExaminerAllowanceRow) -> bool:
    return not (
        (item.bank_name or "").strip()
        and (item.branch_name or "").strip()
        and (item.bank_code or "").strip()
        and (item.account_number or "").strip()
    )


def _grid_border(*, header_bottom: bool = False) -> Border:
    bottom = _SIDE_HEADER_BOTTOM if header_bottom else _SIDE_THIN
    return Border(left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=bottom)


def _write_text(ws: object, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = EXCEL_TEXT_FORMAT


def _write_amount(ws: object, row: int, col: int, value: Decimal) -> None:
    cell = ws.cell(row=row, column=col, value=float(value))
    cell.number_format = AMOUNT_NUMBER_FORMAT


def _build_headers(include_fields: frozenset[str]) -> tuple[list[str], list[int]]:
    headers = list(CORE_HEADER_LABELS)
    widths = list(CORE_COLUMN_WIDTHS)
    # Insert subject names immediately after Subjects (index 4).
    if "subject_names" in include_fields:
        label, width = OPTIONAL_EXPORT_FIELDS["subject_names"]
        headers.insert(5, label)
        widths.insert(5, width)
    # Insert travel zone immediately before T & T (GHS).
    if "travel_zone" in include_fields:
        label, width = OPTIONAL_EXPORT_FIELDS["travel_zone"]
        tt_idx = headers.index("T & T (GHS)")
        headers.insert(tt_idx, label)
        widths.insert(tt_idx, width)
    return headers, widths


def _row_values(item: AdminExaminerAllowanceRow, include_fields: frozenset[str]) -> list[object]:
    incomplete = bank_account_incomplete(item)
    values: list[object] = [
        item.reference_code or "",
        item.full_name,
        _role_label(item.examiner_type),
        item.region,
        item.subject_codes,
    ]
    if "subject_names" in include_fields:
        values.append(item.subject_names or "")
    values.extend(
        [
            item.bank_name or "",
            item.branch_name or "",
            item.bank_code or "",
            item.account_number or "",
            "Incomplete" if incomplete else "OK",
            item.phone_number or "",
            item.responsibility_allowance_ghs,
            item.inconvenience_allowance_ghs,
            item.chief_examiners_report_ghs,
            item.vetting_of_scripts_ghs,
            item.vetting_withholding_tax_ghs,
            item.vetting_net_ghs,
            item.internal_commuting_ghs,
            item.marking_allowance_ghs,
            item.marking_withholding_tax_ghs,
            item.marking_net_ghs,
            item.total_allocated_scripts,
        ]
    )
    if "travel_zone" in include_fields:
        values.append(item.travel_zone_name or "")
    values.extend(
        [
            item.travel_and_transport_ghs,
            item.payout_travel_commuting_ghs,
            item.payout_allowances_marking_ghs,
            item.total_payable_ghs,
        ]
    )
    return values


def detail_workbook_bytes(
    items: list[AdminExaminerAllowanceRow],
    *,
    title: str,
    include_fields: frozenset[str] | None = None,
) -> bytes:
    fields = include_fields or frozenset()
    headers, widths = _build_headers(fields)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Examiner allowances"
    ncols = len(headers)
    start_row = 1

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.fill = _FILL_TITLE
    title_cell.font = _FONT_TITLE
    ws.row_dimensions[start_row].height = 32

    header_row = start_row + 1
    border = _grid_border(header_bottom=True)
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 36

    scripts_col = headers.index("Allocated scripts") + 1
    bank_code_col = headers.index("Bank code") + 1
    account_col = headers.index("Account") + 1
    phone_col = headers.index("Phone") + 1
    region_col = headers.index("Region") + 1
    amount_cols = {
        i
        for i, label in enumerate(headers, start=1)
        if "(GHS)" in label
    }
    text_force_cols = {bank_code_col, account_col, phone_col}
    if "travel_zone" in fields:
        text_force_cols.add(headers.index("Travel zone") + 1)
    if "subject_names" in fields:
        text_force_cols.add(headers.index("Subject names") + 1)

    data_row = header_row + 1
    for idx, item in enumerate(items):
        r = data_row + idx
        values = _row_values(item, fields)
        if bank_account_incomplete(item):
            fill = _FILL_INCOMPLETE
        else:
            fill = _FILL_ZEBRA_ALT if idx % 2 == 1 else _FILL_ZEBRA_BASE
        row_border = _grid_border()
        for col, value in enumerate(values, start=1):
            if col == scripts_col:
                ws.cell(row=r, column=col, value=int(value))
            elif col in amount_cols:
                _write_amount(ws, r, col, Decimal(str(value)))
            elif col in text_force_cols:
                _write_text(ws, r, col, str(value))
            else:
                ws.cell(row=r, column=col, value=value)
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.font = _FONT_DATA
            cell.border = row_border
            if col in amount_cols:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=col == region_col,
                )
        ws.row_dimensions[r].height = 20

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def examiner_detail_workbook_bytes(
    examiners: list[Examiner],
    examination: Examination,
    role_rates: RoleAllowanceMap,
    marking_rates: MarkingRateMap,
    travel_rates: TravelRateMap,
    travel_zones: TravelZoneMap,
    travel_zone_names: TravelZoneNameMap,
    travel_role_factors: TravelRoleFactorMap,
    allocated_booklets: AllocatedBookletsMap,
    source_modes: MarkingScriptSourceModes | None = None,
    *,
    include_fields: frozenset[str] | None = None,
) -> bytes:
    items = examiners_to_admin_rows(
        examiners,
        examination,
        role_rates,
        marking_rates,
        travel_rates,
        travel_zones,
        travel_zone_names,
        travel_role_factors,
        allocated_booklets,
        source_modes,
    )
    title = f"Examiner allowances — {examination_label(examination)}"
    return detail_workbook_bytes(items, title=title, include_fields=include_fields)
