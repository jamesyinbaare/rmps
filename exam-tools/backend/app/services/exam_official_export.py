"""Excel export helpers for exam centre officials."""

import io
import re
import zipfile
from collections import defaultdict
from collections.abc import Sequence
from typing import cast
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import (
    ExamCentreOfficial,
    ExamInspectorSubjectScope,
    ExamOfficialDesignation,
    Examination,
    ExaminationCentre,
    ExaminationDesignationRate,
)
from app.services.exam_official_compensation import compensation_export_values

HEADER_LABELS = [
    "Centre code",
    "Centre name",
    "Full name",
    "Designation",
    "Subject scope",
    "Bank",
    "Branch",
    "Bank code",
    "Account",
    "Days",
    "Phone",
    "Daily rate (GHS)",
    "Commuting per day (GHS)",
    "Airtime (GHS)",
    "Total payable (GHS)",
]

COLUMN_WIDTHS = [12, 28, 26, 18, 22, 26, 12, 16, 6, 14, 14, 14, 14, 14, 16]

EXCEL_TEXT_FORMAT = "@"
DAYS_NUMBER_FORMAT = "0"
DAYS_COLUMN = HEADER_LABELS.index("Days") + 1
PHONE_COLUMN = HEADER_LABELS.index("Phone") + 1

# Exported for tests
HEADER_FILL_COLOR = "2E5077"
TITLE_FILL_COLOR = "1E3A5F"

_FILL_TITLE = PatternFill(fill_type="solid", fgColor=TITLE_FILL_COLOR)
_FILL_PREAMBLE = PatternFill(fill_type="solid", fgColor="F4F6F8")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor=HEADER_FILL_COLOR)
_FILL_ZEBRA_BASE = PatternFill(fill_type="solid", fgColor="FFFFFF")
_FILL_ZEBRA_ALT = PatternFill(fill_type="solid", fgColor="F8FAFC")
_FILL_INVIGILATOR = PatternFill(fill_type="solid", fgColor="ECFDF5")

_FONT_TITLE = Font(bold=True, size=14, color="FFFFFF")
_FONT_PREAMBLE = Font(size=11, color="374151")
_FONT_HEADER = Font(bold=True, size=11, color="FFFFFF")
_FONT_DATA = Font(size=11, color="1F2937")

_SIDE_THIN = Side(style="thin", color="D1D5DB")
_SIDE_HEADER_BOTTOM = Side(style="medium", color="2E5077")


def designation_str(des: object) -> str:
    from app.models import ExamOfficialDesignation

    if isinstance(des, ExamOfficialDesignation):
        return des.value
    return str(des)


def examination_label(ex: Examination) -> str:
    parts = [str(ex.year)]
    if ex.exam_series and str(ex.exam_series).strip():
        parts.append(str(ex.exam_series).strip())
    parts.append(str(ex.exam_type).strip())
    return " ".join(parts)


def safe_filename_part(s: str) -> str:
    t = re.sub(r"[^\w\-]+", "_", s.strip(), flags=re.UNICODE).strip("_")
    return (t or "export")[:80]


def sheet_name_part(s: str) -> str:
    """Sanitize for Excel worksheet tab (keep spaces; drop forbidden characters)."""
    t = s.strip()
    for ch in (":", "\\", "/", "?", "*", "[", "]"):
        t = t.replace(ch, "_")
    return t


def centre_sheet_title(centre: ExaminationCentre, used: set[str]) -> str:
    """Excel sheet name: centre code and name (max 31 chars, unique within workbook)."""
    code = sheet_name_part(cast(str, centre.code)) or "Centre"
    name = sheet_name_part(cast(str, centre.name))
    if name:
        sep = " - "
        room = 31 - len(code) - len(sep)
        base = f"{code}{sep}{name[:room]}" if room >= 1 else code[:31]
    else:
        base = code[:31]
    title = base
    if title in used:
        n = 2
        while True:
            suffix = f"_{n}"
            title = f"{base[: 31 - len(suffix)]}{suffix}"
            if title not in used:
                break
            n += 1
    used.add(title)
    return title


def _grid_border(*, header_bottom: bool = False) -> Border:
    bottom = _SIDE_HEADER_BOTTOM if header_bottom else _SIDE_THIN
    return Border(left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=bottom)


def _header_alignment(column: int) -> Alignment:
    if column in (DAYS_COLUMN, PHONE_COLUMN):
        return Alignment(horizontal="center", vertical="center", wrap_text=False)
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def _data_alignment(column: int) -> Alignment:
    if column == DAYS_COLUMN:
        return Alignment(horizontal="center", vertical="center", wrap_text=False)
    if column == PHONE_COLUMN:
        return Alignment(horizontal="center", vertical="center", wrap_text=False)
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def style_title_row(ws: object, row: int, ncols: int, title: str, *, merge: bool) -> None:
    if merge:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = _FILL_TITLE
    cell.font = _FONT_TITLE
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[row].height = 32


def style_preamble_row(ws: object, row: int, ncols: int, label: str, value: str | int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=f"{label}: {value}")
    cell.fill = _FILL_PREAMBLE
    cell.font = _FONT_PREAMBLE
    cell.number_format = EXCEL_TEXT_FORMAT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20


def style_header_row(ws: object, row: int, ncols: int) -> None:
    border = _grid_border(header_bottom=True)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.border = border
        cell.alignment = _header_alignment(c)
    ws.row_dimensions[row].height = 22


def style_data_row(
    ws: object,
    row: int,
    ncols: int,
    *,
    stripe: bool,
    highlight: bool,
) -> None:
    if highlight:
        fill = _FILL_INVIGILATOR
    elif stripe:
        fill = _FILL_ZEBRA_ALT
    else:
        fill = _FILL_ZEBRA_BASE
    border = _grid_border()
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = _FONT_DATA
        cell.border = border
        cell.alignment = _data_alignment(c)


def write_export_cell(ws: object, row: int, column: int, value: object) -> None:
    """Write a table data cell as Excel text, except the Days column (numeric)."""
    cell = ws.cell(row=row, column=column)
    if column == DAYS_COLUMN:
        cell.value = value
        cell.number_format = DAYS_NUMBER_FORMAT
        return
    cell.number_format = EXCEL_TEXT_FORMAT
    cell.value = "" if value is None else str(value)


def set_col_widths(ws: object, widths: Sequence[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def apply_sheet_finish(ws: object, header_row: int, ncols: int, *, autofilter: bool) -> None:
    ws.freeze_panes = f"A{header_row + 1}"
    if autofilter and ws.max_row >= header_row:
        last_col = get_column_letter(ncols)
        ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"${header_row}:${header_row}"


def data_values(
    off: ExamCentreOfficial,
    centre: ExaminationCentre,
    *,
    rates_by_designation: dict[ExamOfficialDesignation, ExaminationDesignationRate] | None = None,
) -> tuple[str | int, ...]:
    bb = off.bank_branch
    scope = off.subject_scope.value if isinstance(off.subject_scope, ExamInspectorSubjectScope) else str(off.subject_scope)
    rates = rates_by_designation or {}
    comp_vals = compensation_export_values(off, rates)
    return (
        centre.code,
        centre.name,
        off.full_name,
        designation_str(off.designation),
        scope,
        bb.bank_name,
        bb.branch_name,
        str(bb.bank_code),
        off.account_number,
        int(off.num_days),
        off.telephone_number,
        *comp_vals,
    )


def write_centre_block(
    ws: object,
    start_row: int,
    centre: ExaminationCentre,
    exam_label: str,
    pairs: list[tuple[ExamCentreOfficial, ExaminationCentre]],
    *,
    merge_title: bool,
    preamble_rows: list[tuple[str, str | int]] | None = None,
    rates_by_designation: dict[ExamOfficialDesignation, ExaminationDesignationRate] | None = None,
) -> tuple[int, int]:
    """Write optional preamble, title, headers, and data rows; return (next_row, header_row)."""
    ncols = len(HEADER_LABELS)
    r = start_row
    if preamble_rows:
        for label, value in preamble_rows:
            style_preamble_row(ws, r, ncols, label, value)
            r += 1
        r += 1
    title = f"Examination centre: {centre.name} ({centre.code}) · {exam_label}"
    style_title_row(ws, r, ncols, title, merge=merge_title)
    r += 2
    for i, h in enumerate(HEADER_LABELS, start=1):
        ws.cell(row=r, column=i, value=h)
    header_row = r
    style_header_row(ws, r, ncols)
    r += 1
    for data_idx, (off, sch) in enumerate(pairs):
        vals = data_values(off, sch, rates_by_designation=rates_by_designation)
        for i, v in enumerate(vals, start=1):
            write_export_cell(ws, r, i, v)
        is_invigilator = designation_str(off.designation) == "Invigilator"
        style_data_row(ws, r, ncols, stripe=data_idx % 2 == 1, highlight=is_invigilator)
        r += 1
    return r, header_row


def workbook_for_centre(
    centre: ExaminationCentre,
    exam_label: str,
    pairs: list[tuple[ExamCentreOfficial, ExaminationCentre]],
    *,
    preamble_rows: list[tuple[str, str | int]] | None = None,
    rates_by_designation: dict[ExamOfficialDesignation, ExaminationDesignationRate] | None = None,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Officials"
    ncols = len(HEADER_LABELS)
    _, header_row = write_centre_block(
        ws,
        1,
        centre,
        exam_label,
        pairs,
        merge_title=True,
        preamble_rows=preamble_rows,
        rates_by_designation=rates_by_designation,
    )
    set_col_widths(ws, COLUMN_WIDTHS)
    apply_sheet_finish(ws, header_row, ncols, autofilter=True)
    return wb


def workbook_bytes(wb: Workbook) -> bytes:
    xbuf = io.BytesIO()
    wb.save(xbuf)
    return xbuf.getvalue()


def build_zip_export(
    ordered_groups: list[tuple[UUID, list[tuple[ExamCentreOfficial, ExaminationCentre]]]],
    exam_label: str,
    zip_basename: str,
    *,
    rates_by_designation: dict[ExamOfficialDesignation, ExaminationDesignationRate] | None = None,
) -> tuple[bytes, str, str]:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for _cid, plist in ordered_groups:
            centre = plist[0][1]
            wb = workbook_for_centre(centre, exam_label, plist, rates_by_designation=rates_by_designation)
            fname = f"{safe_filename_part(cast(str, centre.code))}-{safe_filename_part(cast(str, centre.name))}.xlsx"
            zf.writestr(fname, workbook_bytes(wb))
    return buf.getvalue(), f"{zip_basename}.zip", "application/zip"


def build_single_sheet_export(
    pairs: list[tuple[ExamCentreOfficial, ExaminationCentre]],
    exam: Examination,
    *,
    sheet_title: str,
    file_base: str,
    rates_by_designation: dict[ExamOfficialDesignation, ExaminationDesignationRate] | None = None,
) -> tuple[bytes, str, str]:
    """One workbook with a single worksheet listing all officials (all centres)."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    tab = sheet_name_part(sheet_title)[:31] or "Officials"
    ws.title = tab
    exam_label = examination_label(exam)
    ncols = len(HEADER_LABELS)
    style_title_row(ws, 1, ncols, f"{sheet_title} · {exam_label}", merge=True)
    r = 3
    for i, h in enumerate(HEADER_LABELS, start=1):
        ws.cell(row=r, column=i, value=h)
    header_row = r
    style_header_row(ws, r, ncols)
    r += 1
    sorted_pairs = sorted(pairs, key=lambda p: (str(p[1].code), str(p[0].full_name)))
    for data_idx, (off, centre) in enumerate(sorted_pairs):
        vals = data_values(off, centre, rates_by_designation=rates_by_designation)
        for i, v in enumerate(vals, start=1):
            write_export_cell(ws, r, i, v)
        is_invigilator = designation_str(off.designation) == "Invigilator"
        style_data_row(ws, r, ncols, stripe=data_idx % 2 == 1, highlight=is_invigilator)
        r += 1
    set_col_widths(ws, COLUMN_WIDTHS)
    apply_sheet_finish(ws, header_row, ncols, autofilter=True)
    return (
        workbook_bytes(wb),
        f"{file_base}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def build_combined_export(
    ordered_groups: list[tuple[UUID, list[tuple[ExamCentreOfficial, ExaminationCentre]]]],
    exam: Examination,
    *,
    rates_by_designation: dict[ExamOfficialDesignation, ExaminationDesignationRate] | None = None,
    file_base: str | None = None,
) -> tuple[bytes, str, str]:
    """Single workbook with one worksheet per examination centre."""
    wb = Workbook()
    exam_label = examination_label(exam)
    ncols = len(HEADER_LABELS)
    used_titles: set[str] = set()
    for idx, (_cid, plist) in enumerate(ordered_groups):
        centre = plist[0][1]
        title = centre_sheet_title(centre, used_titles)
        if idx == 0:
            ws = wb.active
            assert ws is not None
            ws.title = title
        else:
            ws = wb.create_sheet(title=title)
        _, header_row = write_centre_block(
            ws,
            1,
            centre,
            exam_label,
            plist,
            merge_title=True,
            rates_by_designation=rates_by_designation,
        )
        set_col_widths(ws, COLUMN_WIDTHS)
        apply_sheet_finish(ws, header_row, ncols, autofilter=True)
    xbuf = io.BytesIO()
    wb.save(xbuf)
    base = file_base or safe_filename_part(f"exam_{exam.id}_{exam_label}_officials_all_centres")
    return (
        xbuf.getvalue(),
        f"{base}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def group_officials_by_centre(
    pairs: list[tuple[ExamCentreOfficial, ExaminationCentre]],
) -> list[tuple[UUID, list[tuple[ExamCentreOfficial, ExaminationCentre]]]]:
    groups: dict[UUID, list[tuple[ExamCentreOfficial, ExaminationCentre]]] = defaultdict(list)
    for off, centre in pairs:
        groups[centre.id].append((off, centre))
    return sorted(groups.items(), key=lambda kv: kv[1][0][1].code)
