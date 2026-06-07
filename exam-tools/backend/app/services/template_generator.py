"""Excel templates for bulk uploads."""

import io
from datetime import date, time
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from sqlalchemy.ext.asyncio import AsyncSession


def generate_programme_template() -> bytes:
    data = {
        "code": ["PROG01", "PROG02"],
        "name": ["Example Programme 1", "Example Programme 2"],
    }
    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Programmes")
    output.seek(0)
    return output.getvalue()


def generate_subject_template() -> bytes:
    data = {
        "code": ["301", "302", "303", "701"],
        "original_code": ["MATH301", "ENG302", "PHY303", "SCI701"],
        "name": ["Mathematics", "English", "Physics", "Science"],
        "subject_type": ["CORE", "CORE", "CORE", "ELECTIVE"],
        "choice_group_id": ["", "1", "1", ""],
    }
    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Subjects")
    output.seek(0)
    return output.getvalue()


def generate_examiners_bulk_template() -> bytes:
    """Excel template for examiner roster or invitation bulk upload.

    Columns: name, phone_number, subject_code, examiner_type, region.
    Phone column is formatted as Text so Excel does not strip leading zeros.
    """
    df = pd.DataFrame(
        {
            "name": ["Jane Doe", "John Smith"],
            "phone_number": ["0551234567", "0244123456"],
            "subject_code": ["301", "302"],
            "examiner_type": ["assistant_examiner", "chief_examiner"],
            "region": ["Greater Accra", "Ashanti"],
        }
    )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Examiners")
        ws = writer.sheets["Examiners"]
        for row in range(1, 10001):
            ws.cell(row=row, column=2).number_format = "@"
    output.seek(0)
    return output.getvalue()


def generate_inspector_postings_bulk_template() -> bytes:
    """Excel template for inspector postings bulk upload.

    Use center_N + scope_N (ALL/CORE/ELECTIVE) for up to five centres per row.
    Phone column is formatted as Text so Excel does not strip leading zeros on entry.
    """
    df = pd.DataFrame(
        columns=[
            "phone_number",
            "full_name",
            "password",
            "center_1",
            "scope_1",
            "center_2",
            "scope_2",
            "center_3",
            "scope_3",
            "center_4",
            "scope_4",
            "center_5",
            "scope_5",
        ]
    )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Postings")
        ws = writer.sheets["Postings"]
        for row in range(1, 10001):
            ws.cell(row=row, column=1).number_format = "@"
    output.seek(0)
    return output.getvalue()


async def generate_schedule_template(
    session: "AsyncSession",
    exam_year: int | None = None,  # noqa: ARG001
    exam_series: str | None = None,  # noqa: ARG001
    exam_type: str | None = None,  # noqa: ARG001
) -> bytes:
    """
    Excel template for schedule upload: all subjects, plus a Sample Data sheet (registration-portal compatible).
    """
    from sqlalchemy import select

    from app.models import Subject

    stmt = select(Subject).order_by(Subject.code)
    result = await session.execute(stmt)
    subjects = result.scalars().all()

    if subjects:
        data = {
            "subject_code": [
                subject.original_code if subject.original_code else subject.code for subject in subjects
            ],
            "subject_name": [subject.name for subject in subjects],
            "paper1_date": [""] * len(subjects),
            "paper1_start_time": [""] * len(subjects),
            "paper1_end_time": [""] * len(subjects),
            "paper2_date": [""] * len(subjects),
            "paper2_start_time": [""] * len(subjects),
            "paper2_end_time": [""] * len(subjects),
            "write_together": [0] * len(subjects),
        }
    else:
        data = {
            "subject_code": [],
            "subject_name": [],
            "paper1_date": [],
            "paper1_start_time": [],
            "paper1_end_time": [],
            "paper2_date": [],
            "paper2_start_time": [],
            "paper2_end_time": [],
            "write_together": [],
        }

    df = pd.DataFrame(data)

    sample_date = date.today().replace(month=1, day=15)
    sample_data = {
        "subject_code": ["C701", "C702"],
        "subject_name": ["English Language", "Mathematics"],
        "paper1_date": [sample_date, sample_date],
        "paper1_start_time": [time(9, 0), time(9, 0)],
        "paper1_end_time": [time(11, 0), time(11, 30)],
        "paper2_date": [sample_date, sample_date.replace(day=16)],
        "paper2_start_time": [time(9, 0), time(14, 0)],
        "paper2_end_time": [time(11, 0), time(16, 0)],
        "write_together": [1, 0],
    }
    df_sample = pd.DataFrame(sample_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Schedules")
        worksheet = writer.sheets["Schedules"]

        date_columns = ["paper1_date", "paper2_date"]
        time_columns = ["paper1_start_time", "paper1_end_time", "paper2_start_time", "paper2_end_time"]
        number_columns = ["write_together"]

        date_col_letters: list[str] = []
        time_col_letters: list[str] = []
        number_col_letters: list[str] = []

        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            if col_name in date_columns:
                date_col_letters.append(col_letter)
            elif col_name in time_columns:
                time_col_letters.append(col_letter)
            elif col_name in number_columns:
                number_col_letters.append(col_letter)

        for col_letter in date_col_letters:
            for cell in worksheet[col_letter][1:]:
                cell.number_format = "yyyy-mm-dd"

        for col_letter in time_col_letters:
            for cell in worksheet[col_letter][1:]:
                cell.number_format = "HH:mm"

        for col_letter in number_col_letters:
            for cell in worksheet[col_letter][1:]:
                cell.number_format = "0"

        df_sample.to_excel(writer, index=False, sheet_name="Sample Data")
        worksheet_sample = writer.sheets["Sample Data"]

        for col_letter in date_col_letters:
            for cell in worksheet_sample[col_letter][1:]:
                cell.number_format = "yyyy-mm-dd"

        for col_letter in time_col_letters:
            for cell in worksheet_sample[col_letter][1:]:
                cell.number_format = "HH:mm"

        for col_letter in number_col_letters:
            for cell in worksheet_sample[col_letter][1:]:
                cell.number_format = "0"

    output.seek(0)
    return output.getvalue()


def generate_examination_centres_bulk_template(subject_scope: str = "CORE") -> bytes:
    """Excel template for examination centre + membership bulk upload (two columns only)."""
    data = {
        "centre_code": ["H001", "H001"],
        "school_code": ["H001", "S002"],
    }
    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Centres")
    output.seek(0)
    return output.getvalue()


def generate_examiner_invitations_export(rows: list[dict[str, object]]) -> bytes:
    """Excel export of examiner invitations including public URLs."""
    columns = [
        "name",
        "phone_number",
        "subject_code",
        "subject_name",
        "examiner_type",
        "region",
        "status",
        "coordination_date",
        "public_url",
    ]
    df = pd.DataFrame(rows, columns=columns if not rows else None)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Examiner links")
    output.seek(0)
    return output.getvalue()
