"""Parse CSV/XLSX uploads for manual marked script counts."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any
from uuid import UUID

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection

from app.services.examiner_roster import normalize_header_key, read_examiners_spreadsheet
from app.services.school_bulk_upload import inspector_phone_lookup_candidates, parse_inspector_phone_number

TEMPLATE_HEADERS = ("phone_number", "name", "ref_code", "paper", "total_allocation")
_TOTAL_ALLOCATION_COL = 5  # 1-based


def _canonical_column_map() -> dict[str, str]:
    return {
        "phone": "phone_number",
        "mobile": "phone_number",
        "phone_number": "phone_number",
        "total": "total",
        "total_allocation": "total",
        "scripts": "total",
        "count": "total",
        "allocated_scripts": "total",
        "script_count": "total",
    }


def _rename_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [normalize_header_key(c) for c in out.columns]
    cmap = _canonical_column_map()
    rename = {c: cmap[c] for c in out.columns if c in cmap}
    return out.rename(columns=rename)


def read_manual_marked_scripts_spreadsheet(file_bytes: bytes, filename: str) -> pd.DataFrame:
    df = read_examiners_spreadsheet(file_bytes, filename)
    return _rename_dataframe_columns(df)


def _parse_total_cell(raw: Any) -> int:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return 0
    s = str(raw).strip()
    if not s:
        return 0
    if re.fullmatch(r"-?\d+", s):
        value = int(s)
    elif isinstance(raw, (int, float)) and not isinstance(raw, bool):
        if float(raw).is_integer():
            value = int(raw)
        else:
            raise ValueError(f"total must be a whole number, got {raw!r}")
    else:
        raise ValueError(f"total must be a whole number, got {raw!r}")
    if value < 0:
        raise ValueError("total must be >= 0")
    return value


@dataclass
class ManualMarkedScriptsUploadRowError:
    row_number: int
    message: str


@dataclass
class ManualMarkedScriptsUploadResult:
    applied_count: int = 0
    skipped_count: int = 0
    errors: list[ManualMarkedScriptsUploadRowError] = field(default_factory=list)
    items: list[tuple[UUID, int]] = field(default_factory=list)


@dataclass
class ManualMarkedScriptsTemplateRow:
    phone_number: str
    name: str
    ref_code: str
    paper: int
    total_allocation: int | str = ""


def parse_manual_marked_scripts_upload(
    df: pd.DataFrame,
    *,
    phone_to_examiner_id: dict[str, UUID],
) -> ManualMarkedScriptsUploadResult:
    """Parse upload rows; match phones to subject examiners. Does not persist."""
    if "phone_number" not in df.columns:
        raise ValueError("Missing required column: phone_number (aliases: phone, mobile)")
    if "total" not in df.columns:
        raise ValueError(
            "Missing required column: total "
            "(aliases: total_allocation, scripts, count, allocated_scripts)"
        )

    result = ManualMarkedScriptsUploadResult()
    seen_phones: dict[str, int] = {}

    for row_number, (_, row) in enumerate(df.iterrows(), start=2):
        phone_raw = row.get("phone_number")
        if phone_raw is None or (isinstance(phone_raw, float) and pd.isna(phone_raw)):
            result.skipped_count += 1
            continue
        try:
            phone = parse_inspector_phone_number(phone_raw)
        except ValueError as e:
            result.errors.append(ManualMarkedScriptsUploadRowError(row_number=row_number, message=str(e)))
            continue

        if phone in seen_phones:
            result.errors.append(
                ManualMarkedScriptsUploadRowError(
                    row_number=row_number,
                    message=f"Duplicate phone_number {phone!r} (also on row {seen_phones[phone]})",
                )
            )
            continue
        seen_phones[phone] = row_number

        try:
            total = _parse_total_cell(row.get("total"))
        except ValueError as e:
            result.errors.append(ManualMarkedScriptsUploadRowError(row_number=row_number, message=str(e)))
            continue

        examiner_id: UUID | None = None
        for candidate in inspector_phone_lookup_candidates(phone):
            examiner_id = phone_to_examiner_id.get(candidate)
            if examiner_id is not None:
                break
        if examiner_id is None:
            result.errors.append(
                ManualMarkedScriptsUploadRowError(
                    row_number=row_number,
                    message=f"No examiner on this subject matches phone {phone!r}",
                )
            )
            continue

        result.items.append((examiner_id, total))
        if total == 0:
            result.skipped_count += 1
        else:
            result.applied_count += 1

    return result


def generate_manual_marked_scripts_template_bytes(
    *,
    rows: list[ManualMarkedScriptsTemplateRow],
) -> bytes:
    """Build protected XLSX: identity columns locked; only total_allocation editable."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Manual allocation"

    header_font = Font(bold=True)
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.protection = Protection(locked=True)

    for row_idx, row in enumerate(rows, start=2):
        values = (
            row.phone_number,
            row.name,
            row.ref_code,
            row.paper,
            row.total_allocation if row.total_allocation != "" else "",
        )
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == _TOTAL_ALLOCATION_COL:
                # Set alignment before protection — assigning styles later can reset lock flags.
                cell.alignment = Alignment(horizontal="right")
                cell.protection = Protection(locked=False)
            else:
                cell.protection = Protection(locked=True)

    # Extra blank editable rows so users can still type if they insert past the roster.
    extra_end = max(len(rows) + 50, 100)
    for row_idx in range(len(rows) + 2, extra_end + 1):
        for col_idx in range(1, len(TEMPLATE_HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            if col_idx == _TOTAL_ALLOCATION_COL:
                cell.alignment = Alignment(horizontal="right")
                cell.protection = Protection(locked=False)
            else:
                cell.protection = Protection(locked=True)

    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        max_len = len(header)
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 40)

    ws.protection = SheetProtection(
        sheet=True,
        # True = action is blocked. Leave select* False so unlocked total_allocation is editable.
        selectLockedCells=False,
        selectUnlockedCells=False,
        autoFilter=True,
        formatCells=True,
        formatColumns=True,
        formatRows=True,
        insertColumns=True,
        insertRows=True,
        deleteColumns=True,
        deleteRows=True,
        sort=True,
    )

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
