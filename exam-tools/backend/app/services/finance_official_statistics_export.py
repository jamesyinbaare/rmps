"""Excel export for per-centre examination official statistics."""

from __future__ import annotations

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.examination import FinanceCentreOfficialStatisticsRow
from app.schemas.timetable import TimetableDownloadFilter
from app.services.finance_school_summary import subject_filter_filename_suffix

SHEET_NAME = "Official statistics"

HEADER_LABELS = [
    "Centre code",
    "Centre name",
    "Invigilators",
    "Invigilator days",
    "Expected days",
    "Difference",
    "External inspectors",
    "Supervisors",
    "Asst. supervisors",
    "Police",
    "Depot keepers",
    "Total officials",
]

COLUMN_WIDTHS = [14, 32, 12, 14, 14, 12, 16, 12, 16, 10, 14, 14]

COL_VARIANCE = HEADER_LABELS.index("Difference") + 1
NUMERIC_COLS = {3, 4, 5, 6, 7, 8, 9, 10, 11, 12}

_FILL_TITLE = PatternFill(fill_type="solid", fgColor="1E3A5F")
_FILL_PREAMBLE = PatternFill(fill_type="solid", fgColor="F4F6F8")
_FILL_HEADER = PatternFill(fill_type="solid", fgColor="2E5077")
_FILL_ZEBRA_BASE = PatternFill(fill_type="solid", fgColor="FFFFFF")
_FILL_ZEBRA_ALT = PatternFill(fill_type="solid", fgColor="F8FAFC")
_FILL_TOTALS = PatternFill(fill_type="solid", fgColor="EEF2FF")
_FILL_VARIANCE_OVER = PatternFill(fill_type="solid", fgColor="FEE2E2")
_FILL_VARIANCE_SHORT = PatternFill(fill_type="solid", fgColor="FEF3C7")
_FILL_VARIANCE_MATCH = PatternFill(fill_type="solid", fgColor="ECFDF5")

_FONT_TITLE = Font(bold=True, size=14, color="FFFFFF")
_FONT_PREAMBLE = Font(size=11, color="374151")
_FONT_HEADER = Font(bold=True, size=11, color="FFFFFF")
_FONT_DATA = Font(size=11, color="1F2937")
_FONT_TOTALS = Font(bold=True, size=11, color="1F2937")

_SIDE_THIN = Side(style="thin", color="D1D5DB")
_SIDE_HEADER_BOTTOM = Side(style="medium", color="2E5077")


def official_statistics_export_filename(
    exam_label: str,
    subject_filter: TimetableDownloadFilter,
) -> str:
    def part(s: str) -> str:
        t = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", s.strip())
        return (t or "unknown")[:80]

    suffix = subject_filter_filename_suffix(subject_filter)
    return f"{part(exam_label)} official-statistics {suffix}.xlsx"


def _grid_border(*, header_bottom: bool = False) -> Border:
    bottom = _SIDE_HEADER_BOTTOM if header_bottom else _SIDE_THIN
    return Border(left=_SIDE_THIN, right=_SIDE_THIN, top=_SIDE_THIN, bottom=bottom)


def _variance_fill(variance: int) -> PatternFill:
    if variance > 0:
        return _FILL_VARIANCE_OVER
    if variance < 0:
        return _FILL_VARIANCE_SHORT
    return _FILL_VARIANCE_MATCH


def _row_values(row: FinanceCentreOfficialStatisticsRow) -> list[object]:
    return [
        row.center_code,
        row.center_name,
        row.invigilator_count,
        row.invigilator_days,
        row.expected_invigilator_days,
        row.invigilator_variance,
        row.external_inspector,
        row.supervisor,
        row.assistant_supervisor,
        row.police_officer,
        row.depot_keeper,
        row.total_officials,
    ]


def official_statistics_workbook_bytes(
    rows: list[FinanceCentreOfficialStatisticsRow],
    *,
    totals: FinanceCentreOfficialStatisticsRow,
    exam_label: str,
    subject_filter: TimetableDownloadFilter,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME[:31]
    ncols = len(HEADER_LABELS)

    title_row = 1
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=ncols)
    title_cell = ws.cell(row=title_row, column=1, value=f"Official statistics — {exam_label}")
    title_cell.fill = _FILL_TITLE
    title_cell.font = _FONT_TITLE
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[title_row].height = 32

    preamble_row = 2
    scope_label = subject_filter_filename_suffix(subject_filter)
    ws.merge_cells(start_row=preamble_row, start_column=1, end_row=preamble_row, end_column=ncols)
    preamble_cell = ws.cell(row=preamble_row, column=1, value=f"Subject scope: {scope_label}")
    preamble_cell.fill = _FILL_PREAMBLE
    preamble_cell.font = _FONT_PREAMBLE
    preamble_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[preamble_row].height = 20

    header_row = 4
    for col, label in enumerate(HEADER_LABELS, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.border = _grid_border(header_bottom=True)
        cell.alignment = Alignment(
            horizontal="right" if col in NUMERIC_COLS else "left",
            vertical="center",
            wrap_text=True,
        )
    ws.row_dimensions[header_row].height = 36

    data_start = header_row + 1
    all_rows = [*rows, totals]
    for idx, stat_row in enumerate(all_rows):
        excel_row = data_start + idx
        is_totals = stat_row.center_code == "TOTAL"
        stripe = idx % 2 == 1
        values = _row_values(stat_row)
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=value)
            if col == COL_VARIANCE:
                cell.fill = _variance_fill(int(value))
            elif is_totals:
                cell.fill = _FILL_TOTALS
            else:
                cell.fill = _FILL_ZEBRA_ALT if stripe else _FILL_ZEBRA_BASE
            cell.font = _FONT_TOTALS if is_totals else _FONT_DATA
            cell.border = _grid_border()
            cell.alignment = Alignment(
                horizontal="right" if col in NUMERIC_COLS else "left",
                vertical="center",
            )
            if col in NUMERIC_COLS:
                cell.number_format = "0"

    ws.freeze_panes = ws.cell(row=data_start, column=1).coordinate
    last_row = data_start + len(all_rows) - 1
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{last_row}"

    for col, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
