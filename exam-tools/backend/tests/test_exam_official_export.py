"""Excel export layout for exam centre officials."""

import io
from unittest.mock import MagicMock
from uuid import uuid4

from openpyxl import load_workbook

from app.services.exam_official_export import build_combined_export, centre_sheet_title


def test_centre_sheet_title_includes_code_and_name() -> None:
    used: set[str] = set()
    centre = MagicMock()
    centre.code = "ABC01"
    centre.name = "Sample Host School"
    title = centre_sheet_title(centre, used)
    assert title.startswith("ABC01 - ")
    assert "Sample Host School" in title
    assert len(title) <= 31


def test_centre_sheet_title_unique() -> None:
    used: set[str] = set()
    c1 = MagicMock()
    c1.code = "ABC01"
    c1.name = "Host A"
    c2 = MagicMock()
    c2.code = "ABC01"
    c2.name = "Host A"
    t1 = centre_sheet_title(c1, used)
    t2 = centre_sheet_title(c2, used)
    assert t1 != t2
    assert len(t1) <= 31
    assert len(t2) <= 31


def test_build_combined_export_one_worksheet_per_centre() -> None:
    exam = MagicMock()
    exam.id = 1
    exam.year = 2026
    exam.exam_series = None
    exam.exam_type = "BECE"

    def make_pair(code: str) -> tuple[MagicMock, MagicMock]:
        off = MagicMock()
        bb = MagicMock()
        bb.bank_name = "Bank"
        bb.branch_name = "Branch"
        bb.bank_code = "001"
        off.bank_branch = bb
        off.full_name = "Official"
        off.designation = MagicMock(value="Invigilator")
        off.subject_scope = MagicMock(value="CORE")
        off.account_number = "123"
        off.num_days = 1
        off.telephone_number = "0240000000"
        centre = MagicMock()
        centre.id = uuid4()
        centre.code = code
        centre.name = f"Host {code}"
        return off, centre

    groups = [
        (uuid4(), [make_pair("C001")]),
        (uuid4(), [make_pair("C002")]),
    ]
    payload, _filename, _media = build_combined_export(groups, exam)
    wb = load_workbook(io.BytesIO(payload))
    assert wb.sheetnames == ["C001 - Host C001", "C002 - Host C002"]
    assert wb["C001 - Host C001"].max_row >= 4
    assert wb["C002 - Host C002"].max_row >= 4
    assert wb["C001 - Host C001"]["A3"].value == "Centre code"


def test_build_single_sheet_export_one_tab_all_centres() -> None:
    from app.services.exam_official_export import build_single_sheet_export

    exam = MagicMock()
    exam.id = 1
    exam.year = 2026
    exam.exam_series = None
    exam.exam_type = "BECE"

    def make_pair(code: str, designation: str) -> tuple[MagicMock, MagicMock]:
        off = MagicMock()
        bb = MagicMock()
        bb.bank_name = "Bank"
        bb.branch_name = "Branch"
        bb.bank_code = "001"
        off.bank_branch = bb
        off.full_name = f"Person {code}"
        off.designation = MagicMock(value=designation)
        off.subject_scope = MagicMock(value="CORE")
        off.account_number = "123"
        off.num_days = 1
        off.telephone_number = "0240000000"
        centre = MagicMock()
        centre.id = uuid4()
        centre.code = code
        centre.name = f"Host {code}"
        return off, centre

    pairs = [
        make_pair("C001", "Supervisor"),
        make_pair("C002", "Assistant Supervisor"),
    ]
    payload, filename, media = build_single_sheet_export(
        pairs,
        exam,
        sheet_title="Supervisors",
        file_base="exam_1_BECE_supervisors",
    )
    wb = load_workbook(io.BytesIO(payload))
    assert filename == "exam_1_BECE_supervisors.xlsx"
    assert "spreadsheetml.sheet" in media
    assert len(wb.sheetnames) == 1
    ws = wb.active
    assert ws is not None
    assert ws.max_row >= 5
    assert ws["A3"].value == "Centre code"
