"""Unit tests for workforce BoG payment Excel export."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.services.exam_official_bog_export import GRAND_TOTAL_LABEL
from app.services.workforce_payout import (
    DESIGNATION_SCRIPT_CHECKER,
    _bog_rows_from_payout_items,
    workforce_bog_workbook_bytes,
)


def _item(**overrides):
    base = {
        "full_name": "Alice Mensah",
        "phone_number": "0244123456",
        "account_number": "1234567890",
        "bank_code": "030100",
        "completed_scripts": 100,
        "num_days": 3,
        "payable_ghs": Decimal("250.00"),
    }
    base.update(overrides)
    return base


def test_workforce_bog_rows_include_phone_scripts_days_and_incomplete() -> None:
    items = [
        _item(),
        _item(
            full_name="Bob No Bank",
            account_number=None,
            bank_code=None,
            phone_number="0200111222",
            completed_scripts=50,
            num_days=2,
            payable_ghs=Decimal("80.00"),
        ),
        _item(full_name="Carol Idle", completed_scripts=0, payable_ghs=Decimal("0")),
    ]
    rows = _bog_rows_from_payout_items(items, designation=DESIGNATION_SCRIPT_CHECKER)
    by_name = {r.full_name: r for r in rows}

    assert set(by_name) == {"ALICE MENSAH", "BOB NO BANK"}
    assert by_name["ALICE MENSAH"].phone_number == "0244123456"
    assert by_name["ALICE MENSAH"].work_units == 100
    assert by_name["ALICE MENSAH"].num_days == 3
    assert by_name["ALICE MENSAH"].incomplete_bank is False
    assert by_name["ALICE MENSAH"].paper_script_counts == ()
    assert by_name["BOB NO BANK"].incomplete_bank is True
    assert by_name["BOB NO BANK"].account_number == ""
    assert by_name["BOB NO BANK"].work_units == 50


def test_workforce_bog_rows_map_paper_script_counts() -> None:
    items = [
        _item(paper1_script_count=60, paper2_script_count=40, completed_scripts=100),
        _item(
            full_name="Dana Paper2 Only",
            paper1_script_count=0,
            paper2_script_count=25,
            completed_scripts=25,
        ),
    ]
    rows = _bog_rows_from_payout_items(items, designation=DESIGNATION_SCRIPT_CHECKER)
    by_name = {r.full_name: r for r in rows}
    assert by_name["ALICE MENSAH"].paper_script_counts == (60, 40)
    assert by_name["ALICE MENSAH"].work_units == 100
    assert by_name["DANA PAPER2 ONLY"].paper_script_counts == (0, 25)


def test_workforce_bog_workbook_headers_and_highlight() -> None:
    items = [
        _item(),
        _item(
            full_name="Bob No Bank",
            account_number="",
            bank_code="",
            completed_scripts=40,
            num_days=1,
            payable_ghs=Decimal("40.00"),
        ),
    ]
    payload = workforce_bog_workbook_bytes(
        items,
        title="Test SC BoG",
        designation=DESIGNATION_SCRIPT_CHECKER,
        work_unit_label="Scripts",
    )
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws is not None
    headers = [ws.cell(row=2, column=c).value for c in range(1, 10)]
    assert headers == [
        "Serial",
        "Sort code",
        "Account number",
        "Name",
        "Phone",
        "Scripts",
        "Days",
        "Designation",
        "Amount (GHS)",
    ]
    assert ws.cell(row=3, column=4).value == "ALICE MENSAH"
    assert ws.cell(row=3, column=5).value == "0244123456"
    assert ws.cell(row=3, column=6).value == 100
    assert ws.cell(row=3, column=7).value == 3
    # Incomplete (Bob) highlighted amber
    assert ws.cell(row=4, column=1).fill.fgColor.rgb in ("00FDE68A", "FDE68A")
    assert ws.cell(row=5, column=8).value == GRAND_TOTAL_LABEL


def test_workforce_bog_workbook_entries_label() -> None:
    payload = workforce_bog_workbook_bytes(
        [_item(completed_scripts=12, num_days=2, payable_ghs=Decimal("15"))],
        title="Test DEC BoG",
        designation="Data entry clerk",
        work_unit_label="Entries",
    )
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws is not None
    assert ws.cell(row=2, column=6).value == "Entries"
    headers = [ws.cell(row=2, column=c).value for c in range(1, 12) if ws.cell(row=2, column=c).value]
    assert headers == [
        "Serial",
        "Sort code",
        "Account number",
        "Name",
        "Phone",
        "Entries",
        "Days",
        "Designation",
        "Amount (GHS)",
    ]
    assert "Paper 1 scripts" not in headers
    assert "Paper 2 scripts" not in headers


def test_script_checker_bog_workbook_includes_paper_columns() -> None:
    items = [
        _item(
            paper1_script_count=60,
            paper2_script_count=40,
            completed_scripts=100,
            num_days=3,
        ),
        _item(
            full_name="Bob No Bank",
            account_number="",
            bank_code="",
            paper1_script_count=10,
            paper2_script_count=30,
            completed_scripts=40,
            num_days=1,
            payable_ghs=Decimal("40.00"),
        ),
    ]
    payload = workforce_bog_workbook_bytes(
        items,
        title="Test SC BoG",
        designation=DESIGNATION_SCRIPT_CHECKER,
        work_unit_label="Scripts",
        script_paper_numbers=[1, 2],
    )
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws is not None
    headers = [ws.cell(row=2, column=c).value for c in range(1, 13) if ws.cell(row=2, column=c).value]
    assert headers == [
        "Serial",
        "Sort code",
        "Account number",
        "Name",
        "Phone",
        "Paper 1 scripts",
        "Paper 2 scripts",
        "Days",
        "Designation",
        "Amount (GHS)",
    ]
    assert "Scripts" not in headers
    assert ws.cell(row=3, column=4).value == "ALICE MENSAH"
    assert ws.cell(row=3, column=6).value == 60
    assert ws.cell(row=3, column=7).value == 40
    assert ws.cell(row=3, column=8).value == 3
    assert ws.cell(row=4, column=1).fill.fgColor.rgb in ("00FDE68A", "FDE68A")
    assert ws.cell(row=5, column=9).value == GRAND_TOTAL_LABEL
