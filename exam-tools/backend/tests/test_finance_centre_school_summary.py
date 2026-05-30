"""Unit tests for finance centre school summary helpers."""

import io
from datetime import date, datetime, timezone
from unittest.mock import MagicMock
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.models import ExamOfficialDesignation
from app.schemas.examination import (
    FinanceCentreDayInvigilatorRow,
    FinanceCentreInvigilatorSummaryItem,
)
from app.schemas.timetable import TimetableDownloadFilter
from app.services.exam_official_export import (
    DAYS_COLUMN,
    EXCEL_TEXT_FORMAT,
    HEADER_FILL_COLOR,
    HEADER_LABELS,
    workbook_for_centre,
)
from app.services.finance_school_summary import (
    build_role_counts,
    expected_invigilations_total,
    invigilator_days_declared,
    school_summary_export_filename,
    subject_filter_filename_suffix,
)


def _official(designation: ExamOfficialDesignation, num_days: int = 1) -> MagicMock:
    off = MagicMock()
    off.designation = designation
    off.num_days = num_days
    return off


def test_subject_filter_filename_suffix() -> None:
    assert subject_filter_filename_suffix(TimetableDownloadFilter.ALL) == "ALL"
    assert subject_filter_filename_suffix(TimetableDownloadFilter.CORE_ONLY) == "CORE"
    assert subject_filter_filename_suffix(TimetableDownloadFilter.ELECTIVE_ONLY) == "ELECTIVE"


def test_school_summary_export_filename() -> None:
    name = school_summary_export_filename("C001", "Test Centre", TimetableDownloadFilter.CORE_ONLY)
    assert name == "C001 Test Centre CORE.xlsx"


def test_expected_invigilations_total_sums_days() -> None:
    item = FinanceCentreInvigilatorSummaryItem(
        center_id=uuid4(),
        center_code="H1",
        center_name="Host",
        days=[
            FinanceCentreDayInvigilatorRow(
                examination_date=date(2026, 5, 1),
                unique_candidates=30,
                invigilators_required=1,
            ),
            FinanceCentreDayInvigilatorRow(
                examination_date=date(2026, 5, 2),
                unique_candidates=60,
                invigilators_required=2,
            ),
        ],
    )
    assert expected_invigilations_total(item) == 3


def test_build_role_counts() -> None:
    officials = [
        _official(ExamOfficialDesignation.EXTERNAL_INSPECTOR),
        _official(ExamOfficialDesignation.POLICE_OFFICER),
        _official(ExamOfficialDesignation.POLICE_OFFICER),
        _official(ExamOfficialDesignation.SUPERVISOR),
        _official(ExamOfficialDesignation.DEPOT_KEEPER),
        _official(ExamOfficialDesignation.ASSISTANT_SUPERVISOR),
        _official(ExamOfficialDesignation.INVIGILATOR),
    ]
    counts = build_role_counts(officials)
    assert counts.external_inspector == 1
    assert counts.police_officer == 2
    assert counts.supervisor == 1
    assert counts.depot_keeper == 1
    assert counts.assistant_supervisor == 1


def test_invigilator_headcount_in_school_summary() -> None:
    from app.services.finance_school_summary import invigilator_headcount

    officials = [
        _official(ExamOfficialDesignation.INVIGILATOR),
        _official(ExamOfficialDesignation.INVIGILATOR),
        _official(ExamOfficialDesignation.SUPERVISOR),
    ]
    assert invigilator_headcount(officials) == 2


def test_invigilator_days_declared_only_invigilators() -> None:
    officials = [
        _official(ExamOfficialDesignation.INVIGILATOR, 3),
        _official(ExamOfficialDesignation.INVIGILATOR, 2),
        _official(ExamOfficialDesignation.SUPERVISOR, 10),
    ]
    assert invigilator_days_declared(officials) == 5


def test_workbook_includes_preamble_and_official_rows() -> None:
    school = MagicMock()
    school.code = "C100"
    school.name = "Sample School"

    bb = MagicMock()
    bb.bank_name = "Test Bank"
    bb.branch_name = "Main"
    bb.bank_code = "001234"

    off_inv = MagicMock()
    off_inv.full_name = "Jane Doe"
    off_inv.designation = ExamOfficialDesignation.INVIGILATOR
    off_inv.bank_branch = bb
    off_inv.account_number = "1234567890123"
    off_inv.num_days = 4
    off_inv.telephone_number = "0241234567"

    off_sup = MagicMock()
    off_sup.full_name = "Sam Supervisor"
    off_sup.designation = ExamOfficialDesignation.SUPERVISOR
    off_sup.bank_branch = bb
    off_sup.account_number = "9988776655443"
    off_sup.num_days = 2
    off_sup.telephone_number = "0249999999"

    wb = workbook_for_centre(
        school,
        "2026 BECE",
        [(off_inv, school), (off_sup, school)],
        preamble_rows=[
            ("Expected invigilations", 5),
            ("Invigilator days declared", 4),
        ],
    )
    buf = io.BytesIO()
    wb.save(buf)
    loaded = load_workbook(io.BytesIO(buf.getvalue()))
    ws = loaded.active
    assert ws is not None
    values = [ws.cell(row=r, column=1).value for r in range(1, 15)]
    assert "Expected invigilations: 5" in values
    assert "Invigilator days declared: 4" in values
    assert any(v and "Sample School" in str(v) for v in values)
    header_row = next(
        r for r in range(1, 25) if ws.cell(row=r, column=1).value == HEADER_LABELS[0]
    )
    header_fill = ws.cell(row=header_row, column=1).fill.fgColor.rgb
    assert header_fill is not None
    assert str(header_fill).upper().endswith(HEADER_FILL_COLOR)

    assert ws.freeze_panes == f"A{header_row + 1}"
    assert ws.auto_filter.ref is not None

    invigilator_row = next(r for r in range(1, 25) if ws.cell(row=r, column=3).value == "Jane Doe")
    supervisor_row = next(r for r in range(1, 25) if ws.cell(row=r, column=3).value == "Sam Supervisor")
    assert ws.cell(row=invigilator_row, column=9).number_format == EXCEL_TEXT_FORMAT
    assert ws.cell(row=invigilator_row, column=9).value == "1234567890123"
    assert ws.cell(row=invigilator_row, column=DAYS_COLUMN).value == 4
    assert ws.cell(row=invigilator_row, column=DAYS_COLUMN).number_format != EXCEL_TEXT_FORMAT

    invigilator_fill = ws.cell(row=invigilator_row, column=1).fill.fgColor.rgb
    supervisor_fill = ws.cell(row=supervisor_row, column=1).fill.fgColor.rgb
    assert invigilator_fill is not None and supervisor_fill is not None
    assert str(invigilator_fill).upper().endswith("ECFDF5")
    assert str(supervisor_fill).upper().endswith("F8FAFC")
