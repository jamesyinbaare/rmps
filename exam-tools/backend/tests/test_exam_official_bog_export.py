"""Unit tests for BoG payment Excel export."""

from decimal import Decimal
from io import BytesIO
from unittest.mock import MagicMock
from uuid import uuid4

from openpyxl import load_workbook

from app.models import ExamOfficialDesignation, ExaminationDesignationRate
from app.schemas.timetable import TimetableDownloadFilter
from app.services.exam_official_bog_export import (
    GRAND_TOTAL_LABEL,
    bog_export_rows,
    bog_grand_total,
    bog_workbook_bytes,
    centre_bog_export_filename,
)


def _rate(daily: str = "100.00") -> ExaminationDesignationRate:
    rate = MagicMock(spec=ExaminationDesignationRate)
    rate.daily_rate_ghs = Decimal(daily)
    rate.commuting_allowance_ghs = Decimal("0")
    rate.airtime_ghs = Decimal("0")
    return rate


def _official(
    *,
    name: str,
    designation: ExamOfficialDesignation = ExamOfficialDesignation.INVIGILATOR,
    account: str = "1234567890123",
    bank_code: str = "001234",
    num_days: int = 2,
) -> tuple[MagicMock, MagicMock]:
    off = MagicMock()
    off.id = uuid4()
    off.full_name = name
    off.designation = designation
    off.account_number = account
    off.num_days = num_days
    bb = MagicMock()
    bb.bank_code = bank_code
    off.bank_branch = bb
    centre = MagicMock()
    centre.id = uuid4()
    centre.code = "C001"
    centre.name = "Test Centre"
    return off, centre


def test_bog_export_serial_numbers_and_grand_total() -> None:
    rates = {ExamOfficialDesignation.INVIGILATOR: _rate("50.00")}
    pairs = [_official(name="Alice"), _official(name="Bob")]
    rows = bog_export_rows(pairs, rates)
    assert len(rows) == 2
    assert rows[0].serial == "000001"
    assert rows[1].serial == "000002"
    assert bog_grand_total(rows) == Decimal("200.00")


def test_bog_export_excludes_missing_account_and_zero_payable() -> None:
    rates = {ExamOfficialDesignation.INVIGILATOR: _rate("50.00")}
    pairs = [
        _official(name="Alice", account=""),
        _official(name="Bob", num_days=0),
        _official(name="Carol"),
    ]
    rows = bog_export_rows(pairs, rates)
    assert len(rows) == 1
    assert rows[0].full_name == "CAROL"


def test_bog_workbook_text_columns_and_total_row() -> None:
    rates = {
        ExamOfficialDesignation.INVIGILATOR: _rate("10.00"),
        ExamOfficialDesignation.SUPERVISOR: _rate("20.00"),
    }
    pairs = [
        _official(name="Supervisor One", designation=ExamOfficialDesignation.SUPERVISOR, bank_code="000012"),
        _official(name="Invigilator One", bank_code="001234"),
    ]
    payload = bog_workbook_bytes(pairs, rates, title="Test export")
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws is not None
    assert ws.cell(row=2, column=1).value == "Serial"
    assert ws.cell(row=3, column=4).value == "SUPERVISOR ONE"
    assert ws.cell(row=3, column=1).value == "000001"
    assert ws.cell(row=3, column=2).value == "000012"
    assert ws.cell(row=3, column=2).number_format == "@"
    assert ws.cell(row=3, column=3).number_format == "@"
    assert ws.cell(row=4, column=1).value == "000002"
    total_row = 5
    assert ws.cell(row=total_row, column=5).value == GRAND_TOTAL_LABEL
    assert ws.cell(row=total_row, column=6).value == 60.0


def test_centre_bog_export_filename() -> None:
    name = centre_bog_export_filename("C001", "Test Centre", TimetableDownloadFilter.CORE_ONLY)
    assert name == "C001 Test Centre BoG CORE.xlsx"
