"""Unit tests for examiner BoG payment Excel export."""

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook

from app.schemas.admin_examiner_allowance import AdminExaminerAllowanceRow
from app.schemas.examination_examiner_allowance_rate import SubjectMarkingBreakdownRow
from app.services.exam_official_bog_export import GRAND_TOTAL_LABEL
from app.services.exam_official_bog_export import bog_workbook_bytes
from app.services.examiner_allowance_bog_export import bog_rows_from_admin_items


def _row(
    *,
    name: str,
    role: str = "assistant_examiner",
    account: str = "1234567890123",
    bank_code: str = "001234",
    total: str = "200.00",
) -> AdminExaminerAllowanceRow:
    return AdminExaminerAllowanceRow(
        id=uuid4(),
        examination_id=1,
        examination_label="BECE 2026",
        full_name=name,
        examiner_type=role,
        region="Ashanti",
        subject_codes="MATH",
        subject_names="Mathematics",
        bank_branch_id=uuid4(),
        bank_code=bank_code,
        bank_name="GCB",
        branch_name="Kumasi",
        account_number=account,
        phone_number="0550000000",
        responsibility_allowance_ghs=Decimal(total),
        inconvenience_allowance_ghs=Decimal("0"),
        chief_examiners_report_ghs=Decimal("0"),
        vetting_of_scripts_ghs=Decimal("0"),
        internal_commuting_ghs=Decimal("0"),
        marking_allowance_ghs=Decimal("0"),
        travel_base_ghs=Decimal("0"),
        travel_zone_name=None,
        travel_role_factor=Decimal("1"),
        travel_and_transport_ghs=Decimal("0"),
        total_allocated_scripts=0,
        marking_withholding_tax_ghs=Decimal("0"),
        marking_net_ghs=Decimal("0"),
        vetting_withholding_tax_ghs=Decimal("0"),
        vetting_net_ghs=Decimal("0"),
        payout_travel_commuting_ghs=Decimal("0"),
        payout_allowances_marking_ghs=Decimal(total),
        total_payable_ghs=Decimal(total),
        subject_breakdowns=[
            SubjectMarkingBreakdownRow(
                subject_id=1,
                subject_code="MATH",
                subject_name="Mathematics",
                paper_number=1,
                allocated_booklets=0,
                marking_allowance_ghs=Decimal("0"),
            )
        ],
        created_at=datetime(2026, 6, 1),
        updated_at=datetime(2026, 6, 1),
    )


def test_bog_rows_include_missing_bank_and_zero_total() -> None:
    items = [
        _row(name="Alice", account=""),
        _row(name="Bob", total="0"),
        _row(name="Carol"),
    ]
    rows = bog_rows_from_admin_items(items)
    assert len(rows) == 3
    by_name = {r.full_name: r for r in rows}
    assert by_name["ALICE"].account_number == ""
    assert by_name["ALICE"].sort_code == "001234"
    assert by_name["ALICE"].phone_number == "0550000000"
    assert by_name["ALICE"].incomplete_bank is True
    assert by_name["BOB"].amount == Decimal("0")
    assert by_name["CAROL"].account_number == "1234567890123"
    assert by_name["CAROL"].amount == Decimal("200.00")
    assert by_name["CAROL"].incomplete_bank is False


def test_examiner_bog_workbook_grand_total() -> None:
    items = [_row(name="Alice", total="100"), _row(name="Bob", total="50")]
    rows = bog_rows_from_admin_items(items)
    payload = bog_workbook_bytes([], {}, title="Test", prebuilt_rows=rows)
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws is not None
    assert ws.cell(row=5, column=5).value == GRAND_TOTAL_LABEL
    assert ws.cell(row=5, column=6).value == 150.0


def test_bog_rows_use_travel_commuting_bucket() -> None:
    item = _row(name="Alice", total="200.00")
    item = item.model_copy(
        update={
            "payout_travel_commuting_ghs": Decimal("50.00"),
            "payout_allowances_marking_ghs": Decimal("150.00"),
            "total_payable_ghs": Decimal("200.00"),
        }
    )
    from app.services.examiner_allowance_bog_export import ExaminerBogPayoutMode, bog_rows_from_admin_items

    rows = bog_rows_from_admin_items([item], ExaminerBogPayoutMode.TRAVEL_COMMUTING)
    assert len(rows) == 1
    assert rows[0].amount == Decimal("50.00")


def test_bog_rows_use_allowances_marking_bucket() -> None:
    item = _row(name="Alice", total="200.00")
    item = item.model_copy(
        update={
            "payout_travel_commuting_ghs": Decimal("50.00"),
            "payout_allowances_marking_ghs": Decimal("150.00"),
            "total_payable_ghs": Decimal("200.00"),
        }
    )
    from app.services.examiner_allowance_bog_export import ExaminerBogPayoutMode, bog_rows_from_admin_items

    rows = bog_rows_from_admin_items([item], ExaminerBogPayoutMode.ALLOWANCES_MARKING)
    assert len(rows) == 1
    assert rows[0].amount == Decimal("150.00")


def _row_with_papers(*, name: str = "Alice") -> AdminExaminerAllowanceRow:
    return _row(name=name).model_copy(
        update={
            "total_allocated_scripts": 30,
            "subject_breakdowns": [
                SubjectMarkingBreakdownRow(
                    subject_id=1,
                    subject_code="MATH",
                    subject_name="Mathematics",
                    paper_number=1,
                    allocated_booklets=10,
                    marking_allowance_ghs=Decimal("0"),
                ),
                SubjectMarkingBreakdownRow(
                    subject_id=1,
                    subject_code="MATH",
                    subject_name="Mathematics",
                    paper_number=2,
                    allocated_booklets=20,
                    marking_allowance_ghs=Decimal("0"),
                ),
            ],
        }
    )


def test_bog_all_and_allowances_include_script_paper_columns() -> None:
    from app.services.examiner_allowance_bog_export import ExaminerBogPayoutMode, bog_rows_from_admin_items

    item = _row_with_papers()
    for mode in (ExaminerBogPayoutMode.ALL, ExaminerBogPayoutMode.ALLOWANCES_MARKING):
        rows = bog_rows_from_admin_items(
            [item],
            mode,
            subject_id=1,
            paper_numbers=[1, 2],
        )
        payload = bog_workbook_bytes(
            [],
            {},
            title="Test",
            prebuilt_rows=rows,
            include_phone=True,
            script_paper_numbers=[1, 2],
        )
        wb = load_workbook(BytesIO(payload))
        ws = wb.active
        assert ws is not None
        headers = [ws.cell(row=2, column=c).value for c in range(1, 16) if ws.cell(row=2, column=c).value]
        assert "Allocated scripts" in headers
        assert "Paper 1 scripts" in headers
        assert "Paper 2 scripts" in headers
        assert headers.index("Designation") < headers.index("Allocated scripts")
        assert headers.index("Paper 2 scripts") < headers.index("Amount (GHS)")
        alloc_col = headers.index("Allocated scripts") + 1
        p1_col = headers.index("Paper 1 scripts") + 1
        p2_col = headers.index("Paper 2 scripts") + 1
        assert ws.cell(row=3, column=alloc_col).value == 30
        assert ws.cell(row=3, column=p1_col).value == 10
        assert ws.cell(row=3, column=p2_col).value == 20


def test_bog_travel_mode_omits_script_paper_columns() -> None:
    from app.services.examiner_allowance_bog_export import ExaminerBogPayoutMode, bog_rows_from_admin_items

    item = _row_with_papers()
    rows = bog_rows_from_admin_items(
        [item],
        ExaminerBogPayoutMode.TRAVEL_COMMUTING,
        subject_id=1,
        paper_numbers=[1, 2],
    )
    assert rows[0].allocated_scripts is None
    assert rows[0].paper_script_counts == ()
    payload = bog_workbook_bytes(
        [],
        {},
        title="Test",
        prebuilt_rows=rows,
        include_phone=True,
        script_paper_numbers=None,
    )
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws is not None
    headers = [ws.cell(row=2, column=c).value for c in range(1, 16) if ws.cell(row=2, column=c).value]
    assert "Allocated scripts" not in headers
    assert "Paper 1 scripts" not in headers
    assert headers == [
        "Serial",
        "Sort code",
        "Account number",
        "Name",
        "Reference code",
        "Phone",
        "Designation",
        "Amount (GHS)",
    ]
