"""Unit tests for finance centre inspector analysis."""

import io
from decimal import Decimal
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.models import ExamOfficialDesignation
from app.schemas.examination import (
    FinanceCentreDayInvigilatorRow,
    FinanceCentreInspectorAnalysisRow,
    FinanceCentreInvigilatorSummaryItem,
)
from app.schemas.timetable import TimetableDownloadFilter
from app.services.finance_inspector_analysis import (
    DEFAULT_INSPECTOR_CANDIDATES_RATIO,
    TOTALS_ROW_ID,
    build_inspector_analysis_row,
    external_inspector_max_assigned_days,
    external_inspector_pay_at_exam_days,
    external_inspector_pay_total,
    inspector_phone_dedup_key,
    inspectors_required_headcount,
    pay_at_posted_headcount,
    sum_inspector_analysis_rows,
    unique_phones_from_paid_inspectors,
    unique_phones_from_posted_inspectors,
)
from app.services.finance_inspector_analysis_export import (
    HEADER_LABELS,
    inspector_analysis_export_filename,
    inspector_analysis_workbook_bytes,
)


def _official(
    phone: str,
    *,
    designation=ExamOfficialDesignation.EXTERNAL_INSPECTOR,
    num_days: int = 2,
) -> MagicMock:
    off = MagicMock()
    off.id = uuid4()
    off.designation = designation
    off.telephone_number = phone
    off.num_days = num_days
    return off


def _rate(*, daily: str = "100", commuting: str = "10", airtime: str = "50") -> MagicMock:
    rate = MagicMock()
    rate.daily_rate_ghs = Decimal(daily)
    rate.commuting_allowance_ghs = Decimal(commuting)
    rate.airtime_ghs = Decimal(airtime)
    return rate


def _row_kwargs(**overrides: object) -> dict[str, object]:
    base: dict[str, object] = {
        "total_candidates": 600,
        "exam_days": 2,
        "paid_phones": set(),
        "posted_phones": set(),
        "total_inspector_pay_ghs": Decimal("0"),
        "max_inspector_assigned_days": 0,
        "pay_at_exam_days_ghs": Decimal("0"),
        "pay_at_assigned_days_ghs": Decimal("0"),
        "pay_at_posted_count_ghs": Decimal("0"),
    }
    base.update(overrides)
    return base


def _centre(code: str = "C001", name: str = "Test Centre") -> MagicMock:
    centre = MagicMock()
    centre.id = uuid4()
    centre.code = code
    centre.name = name
    centre.region = None
    return centre


def _posting_user(phone: str | None) -> tuple[MagicMock, MagicMock]:
    posting = MagicMock()
    posting.inspector_user_id = uuid4()
    posting.subject_scope = "ALL"
    user = MagicMock()
    user.phone_number = phone
    return posting, user


def test_inspectors_required_headcount() -> None:
    assert inspectors_required_headcount(0) == 0
    assert inspectors_required_headcount(300) == 1
    assert inspectors_required_headcount(301) == 2
    assert inspectors_required_headcount(600, ratio=300) == 2
    assert inspectors_required_headcount(600, ratio=200) == 3
    assert inspectors_required_headcount(DEFAULT_INSPECTOR_CANDIDATES_RATIO, ratio=300) == 1


def test_build_inspector_analysis_row_custom_ratio() -> None:
    centre = _centre()
    row = build_inspector_analysis_row(
        centre,
        subject_filter=TimetableDownloadFilter.ALL,
        candidates_per_inspector=200,
        **_row_kwargs(total_candidates=600),
    )
    assert row.inspectors_required == 3


def test_phone_dedup_normalizes_ghana_formats() -> None:
    assert inspector_phone_dedup_key("0551234567", fallback="x") == inspector_phone_dedup_key(
        "233551234567", fallback="y"
    )


def test_unique_phones_from_paid_inspectors_dedupes() -> None:
    officials = [
        _official("0551111111"),
        _official("0551111111"),
        _official("0552222222"),
        _official("0553333333", designation=ExamOfficialDesignation.SUPERVISOR),
    ]
    phones = unique_phones_from_paid_inspectors(officials)
    assert len(phones) == 2


def test_unique_phones_from_posted_inspectors_dedupes() -> None:
    pairs = [
        _posting_user("0554444444"),
        _posting_user("233554444444"),
        _posting_user("0555555555"),
    ]
    phones = unique_phones_from_posted_inspectors(pairs, subject_filter=TimetableDownloadFilter.ALL)
    assert len(phones) == 2


def test_build_inspector_analysis_row_overlap_and_variance() -> None:
    centre = _centre()
    paid = {inspector_phone_dedup_key("0551111111", fallback="a")}
    posted = {
        inspector_phone_dedup_key("0551111111", fallback="b"),
        inspector_phone_dedup_key("0552222222", fallback="c"),
    }
    row = build_inspector_analysis_row(
        centre,
        subject_filter=TimetableDownloadFilter.ALL,
        **_row_kwargs(
            exam_days=3,
            paid_phones=paid,
            posted_phones=posted,
            total_inspector_pay_ghs=Decimal("1500.00"),
            pay_at_assigned_days_ghs=Decimal("1500.00"),
        ),
    )
    assert row.inspectors_required == 2
    assert row.external_inspector_count == 1
    assert row.posted_inspector_count == 2
    assert row.unique_inspector_count == 2
    assert row.inspectors_in_both == 1
    assert row.paid_inspector_variance == -1
    assert row.candidates_per_paid_inspector == 600.0
    assert row.total_inspector_pay_ghs == Decimal("1500.00")


def test_external_inspector_max_assigned_days() -> None:
    assert external_inspector_max_assigned_days([]) == 0
    officials = [_official("0551111111", num_days=3), _official("0552222222", num_days=5)]
    assert external_inspector_max_assigned_days(officials) == 5
    mixed = officials + [_official("0553333333", designation=ExamOfficialDesignation.SUPERVISOR, num_days=99)]
    assert external_inspector_max_assigned_days(mixed) == 5


def test_external_inspector_pay_at_exam_days_vs_assigned() -> None:
    officials = [_official("0551111111", num_days=3), _official("0552222222", num_days=5)]
    rates = {ExamOfficialDesignation.EXTERNAL_INSPECTOR: _rate()}
    exam_days = 2
    # per official at exam_days: 100*2 + 10*2 + 50 = 270; two officials = 540
    assert external_inspector_pay_at_exam_days(officials, rates, exam_days) == Decimal("540")
    # assigned: (100*3+10*3+50) + (100*5+10*5+50) = 380 + 600 = 980
    assert external_inspector_pay_total(officials, rates) == Decimal("980")


def test_pay_at_posted_headcount() -> None:
    rates = {ExamOfficialDesignation.EXTERNAL_INSPECTOR: _rate()}
    # single slot at 2 exam days: 270; 3 posted = 810
    assert pay_at_posted_headcount(3, 2, rates) == Decimal("810")
    assert pay_at_posted_headcount(0, 2, rates) == Decimal("0")


def test_build_inspector_analysis_row_days_and_pay_variances() -> None:
    centre = _centre()
    row = build_inspector_analysis_row(
        centre,
        subject_filter=TimetableDownloadFilter.ALL,
        **_row_kwargs(
            exam_days=2,
            max_inspector_assigned_days=5,
            pay_at_exam_days_ghs=Decimal("540"),
            pay_at_assigned_days_ghs=Decimal("980"),
            total_inspector_pay_ghs=Decimal("980"),
            pay_at_posted_count_ghs=Decimal("810"),
        ),
    )
    assert row.assigned_days_variance == 3
    assert row.days_pay_variance_ghs == Decimal("440")
    assert row.payroll_vs_posted_variance_ghs == Decimal("170")


def test_external_inspector_pay_total() -> None:
    officials = [_official("0551111111"), _official("0552222222")]
    rate = MagicMock()
    rate.daily_rate_ghs = Decimal("100")
    rate.commuting_allowance_ghs = Decimal("0")
    rate.airtime_ghs = Decimal("0")
    rates = {ExamOfficialDesignation.EXTERNAL_INSPECTOR: rate}
    total = external_inspector_pay_total(officials, rates)
    assert total == Decimal("400")


def test_sum_inspector_analysis_rows() -> None:
    id1, id2 = uuid4(), uuid4()
    rows = [
        FinanceCentreInspectorAnalysisRow(
            center_id=id1,
            center_code="A",
            center_name="Centre A",
            subject_filter="ALL",
            total_candidates=300,
            exam_days=2,
            external_inspector_count=1,
            posted_inspector_count=2,
            unique_inspector_count=2,
            inspectors_in_both=1,
            total_inspector_pay_ghs=Decimal("500"),
            max_inspector_assigned_days=2,
            assigned_days_variance=0,
            pay_at_exam_days_ghs=Decimal("400"),
            pay_at_assigned_days_ghs=Decimal("500"),
            days_pay_variance_ghs=Decimal("100"),
            pay_at_posted_count_ghs=Decimal("300"),
            payroll_vs_posted_variance_ghs=Decimal("200"),
            inspectors_required=1,
            paid_inspector_variance=0,
            candidates_per_paid_inspector=300.0,
        ),
        FinanceCentreInspectorAnalysisRow(
            center_id=id2,
            center_code="B",
            center_name="Centre B",
            subject_filter="ALL",
            total_candidates=600,
            exam_days=4,
            external_inspector_count=3,
            posted_inspector_count=1,
            unique_inspector_count=3,
            inspectors_in_both=1,
            total_inspector_pay_ghs=Decimal("900"),
            max_inspector_assigned_days=4,
            assigned_days_variance=0,
            pay_at_exam_days_ghs=Decimal("800"),
            pay_at_assigned_days_ghs=Decimal("900"),
            days_pay_variance_ghs=Decimal("100"),
            pay_at_posted_count_ghs=Decimal("200"),
            payroll_vs_posted_variance_ghs=Decimal("700"),
            inspectors_required=2,
            paid_inspector_variance=1,
            candidates_per_paid_inspector=200.0,
        ),
    ]
    totals = sum_inspector_analysis_rows(rows)
    assert totals.center_id == TOTALS_ROW_ID
    assert totals.total_candidates == 900
    assert totals.external_inspector_count == 4
    assert totals.inspectors_required == 3
    assert totals.paid_inspector_variance == 1
    assert totals.total_inspector_pay_ghs == Decimal("1400")
    assert totals.max_inspector_assigned_days == 6
    assert totals.assigned_days_variance == 0
    assert totals.pay_at_exam_days_ghs == Decimal("1200")
    assert totals.pay_at_assigned_days_ghs == Decimal("1400")
    assert totals.days_pay_variance_ghs == Decimal("200")
    assert totals.pay_at_posted_count_ghs == Decimal("500")
    assert totals.payroll_vs_posted_variance_ghs == Decimal("900")


def test_inspector_analysis_export_filename() -> None:
    name = inspector_analysis_export_filename("2026 MAY — BECE", TimetableDownloadFilter.CORE_ONLY)
    assert name == "2026 MAY — BECE inspector-analysis CORE full.xlsx"
    staffing = inspector_analysis_export_filename(
        "2026 MAY — BECE", TimetableDownloadFilter.CORE_ONLY, export_variant="staffing"
    )
    assert staffing == "2026 MAY — BECE inspector-analysis CORE staffing.xlsx"
    pay = inspector_analysis_export_filename(
        "2026 MAY — BECE", TimetableDownloadFilter.CORE_ONLY, export_variant="pay_variance"
    )
    assert pay == "2026 MAY — BECE inspector-analysis CORE pay-variance.xlsx"
    rich = inspector_analysis_export_filename(
        "2026 MAY — BECE", TimetableDownloadFilter.CORE_ONLY, export_variant="staffing", export_style="rich"
    )
    assert rich == "2026 MAY — BECE inspector-analysis CORE staffing formatted.xlsx"


def test_inspector_analysis_rich_workbook_has_legend_and_groups() -> None:
    centre_id = uuid4()
    rows = [
        FinanceCentreInspectorAnalysisRow(
            center_id=centre_id,
            center_code="C001",
            center_name="Host",
            subject_filter="ALL",
            total_candidates=300,
            exam_days=2,
            external_inspector_count=2,
            posted_inspector_count=3,
            unique_inspector_count=3,
            inspectors_in_both=1,
            inspectors_required=1,
            paid_inspector_variance=1,
            total_inspector_pay_ghs=Decimal("250"),
            max_inspector_assigned_days=2,
            assigned_days_variance=0,
            pay_at_exam_days_ghs=Decimal("200"),
            pay_at_assigned_days_ghs=Decimal("250"),
            days_pay_variance_ghs=Decimal("50"),
            pay_at_posted_count_ghs=Decimal("0"),
            payroll_vs_posted_variance_ghs=Decimal("250"),
        ),
    ]
    totals = sum_inspector_analysis_rows(rows)
    payload = inspector_analysis_workbook_bytes(
        rows,
        totals=totals,
        exam_label="2026 BECE",
        subject_filter=TimetableDownloadFilter.ALL,
        candidates_per_inspector=250,
        export_variant="staffing",
        export_style="rich",
    )
    wb = load_workbook(io.BytesIO(payload))
    assert "Legend" in wb.sheetnames
    ws = wb["Inspector analysis"]
    assert ws.cell(row=1, column=1).value.endswith("(Staffing — formatted)")
    assert "Rule:" in str(ws.cell(row=4, column=1).value)
    assert ws.cell(row=6, column=1).value == "Centre"
    assert ws.cell(row=7, column=1).value == "Centre code"
    assert ws.cell(row=8, column=10).value == 2


def test_inspector_analysis_workbook_bytes() -> None:
    centre_id = uuid4()
    rows = [
        FinanceCentreInspectorAnalysisRow(
            center_id=centre_id,
            center_code="C001",
            center_name="Host",
            subject_filter="ALL",
            total_candidates=300,
            exam_days=2,
            external_inspector_count=1,
            inspectors_required=1,
            paid_inspector_variance=0,
            total_inspector_pay_ghs=Decimal("250"),
            max_inspector_assigned_days=2,
            assigned_days_variance=0,
            pay_at_exam_days_ghs=Decimal("200"),
            pay_at_assigned_days_ghs=Decimal("250"),
            days_pay_variance_ghs=Decimal("50"),
            pay_at_posted_count_ghs=Decimal("0"),
            payroll_vs_posted_variance_ghs=Decimal("250"),
        ),
    ]
    totals = sum_inspector_analysis_rows(rows)
    payload = inspector_analysis_workbook_bytes(
        rows,
        totals=totals,
        exam_label="2026 BECE",
        subject_filter=TimetableDownloadFilter.ALL,
        candidates_per_inspector=250,
    )
    wb = load_workbook(io.BytesIO(payload))
    ws = wb["Inspector analysis"]
    assert ws.cell(row=1, column=1).value == "Inspector analysis — 2026 BECE (Full report)"
    assert ws.cell(row=3, column=1).value == "Ratio: 250 candidates per inspector"
    assert ws.cell(row=5, column=1).value == HEADER_LABELS[0]
    assert ws.cell(row=5, column=5).value == "Max assigned days"
    assert ws.cell(row=6, column=1).value == "C001"
    assert ws.cell(row=7, column=1).value == "TOTAL"


@pytest.mark.asyncio
async def test_build_finance_inspector_analysis_aggregates_centres() -> None:
    centre = _centre()
    session = AsyncMock()
    loaded_row = build_inspector_analysis_row(
        centre,
        subject_filter=TimetableDownloadFilter.ALL,
        **_row_kwargs(total_candidates=300, exam_days=1),
    )
    inv_item = FinanceCentreInvigilatorSummaryItem(
        center_id=centre.id,
        center_code="C001",
        center_name="Test Centre",
        days=[FinanceCentreDayInvigilatorRow(examination_date=__import__("datetime").date(2026, 5, 1), unique_candidates=300, invigilators_required=10)],
    )

    with (
        patch(
            "app.services.finance_inspector_analysis.list_centres_for_official_statistics",
            new_callable=AsyncMock,
        ) as mock_centres,
        patch(
            "app.services.finance_inspector_analysis.build_inspector_analysis_row_for_centre",
            new_callable=AsyncMock,
        ) as mock_row,
        patch(
            "app.services.exam_official_compensation.load_designation_rates_map",
            new_callable=AsyncMock,
        ) as mock_rates,
    ):
        mock_centres.return_value = [centre]
        mock_row.return_value = loaded_row
        mock_rates.return_value = {}

        from app.services.finance_inspector_analysis import build_finance_inspector_analysis

        build_invigilator = AsyncMock(return_value=inv_item)
        response = await build_finance_inspector_analysis(
            session,
            1,
            TimetableDownloadFilter.ALL,
            build_invigilator_item=build_invigilator,
            candidates_per_inspector=250,
        )

    assert len(response.centres) == 1
    assert response.centres[0].center_code == "C001"
    assert response.totals.center_code == "TOTAL"
    assert response.candidates_per_inspector == 250
