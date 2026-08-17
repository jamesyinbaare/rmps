"""Tests for manual marked scripts upload parsing and template generation."""

from io import BytesIO
from uuid import uuid4

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.services.manual_marked_scripts_upload import (
    ManualMarkedScriptsTemplateRow,
    TEMPLATE_HEADERS,
    generate_manual_marked_scripts_template_bytes,
    parse_manual_marked_scripts_upload,
    read_manual_marked_scripts_spreadsheet,
)


def test_upload_matches_phone_and_applies_count() -> None:
    examiner_id = uuid4()
    phone_map = {"244123456": examiner_id, "244123456.0": examiner_id}
    df = pd.DataFrame([{"phone_number": "244123456", "total": 25}])
    result = parse_manual_marked_scripts_upload(df, phone_to_examiner_id=phone_map)
    assert result.errors == []
    assert result.applied_count == 1
    assert result.items == [(examiner_id, 25)]


def test_upload_accepts_total_allocation_alias() -> None:
    examiner_id = uuid4()
    phone_map = {"244123456": examiner_id}
    df = pd.DataFrame([{"phone_number": "244123456", "total_allocation": 18, "name": "Ada", "paper": 1}])
    # Simulate spreadsheet header normalisation path
    raw = BytesIO()
    df.to_excel(raw, index=False, engine="openpyxl")
    parsed_df = read_manual_marked_scripts_spreadsheet(raw.getvalue(), "upload.xlsx")
    result = parse_manual_marked_scripts_upload(parsed_df, phone_to_examiner_id=phone_map)
    assert result.errors == []
    assert result.applied_count == 1
    assert result.items == [(examiner_id, 18)]


def test_upload_unknown_phone_error() -> None:
    df = pd.DataFrame([{"phone_number": "999000111", "total": 10}])
    result = parse_manual_marked_scripts_upload(df, phone_to_examiner_id={})
    assert len(result.errors) == 1
    assert "No examiner" in result.errors[0].message
    assert result.items == []


def test_upload_duplicate_phone_error() -> None:
    examiner_id = uuid4()
    phone_map = {"244123456": examiner_id}
    df = pd.DataFrame(
        [
            {"phone_number": "244123456", "total": 10},
            {"phone_number": "244123456", "total": 12},
        ]
    )
    result = parse_manual_marked_scripts_upload(df, phone_to_examiner_id=phone_map)
    assert any("Duplicate phone_number" in e.message for e in result.errors)


def test_upload_blank_total_skipped() -> None:
    examiner_id = uuid4()
    phone_map = {"244123456": examiner_id}
    df = pd.DataFrame([{"phone_number": "244123456", "total": ""}])
    result = parse_manual_marked_scripts_upload(df, phone_to_examiner_id=phone_map)
    assert result.errors == []
    assert result.skipped_count == 1
    assert result.items == [(examiner_id, 0)]


def test_upload_missing_columns_raises() -> None:
    df = pd.DataFrame([{"phone_number": "244123456"}])
    with pytest.raises(ValueError, match="total"):
        parse_manual_marked_scripts_upload(df, phone_to_examiner_id={})


def test_template_includes_examiners_and_locks_identity_columns() -> None:
    rows = [
        ManualMarkedScriptsTemplateRow(
            phone_number="244111222",
            name="Ada Lovelace",
            ref_code="MATH-ADA1",
            paper=2,
            total_allocation=40,
        ),
        ManualMarkedScriptsTemplateRow(
            phone_number="244333444",
            name="Grace Hopper",
            ref_code="",
            paper=2,
            total_allocation="",
        ),
    ]
    content = generate_manual_marked_scripts_template_bytes(rows=rows)
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, len(TEMPLATE_HEADERS) + 1)]
    assert tuple(headers) == TEMPLATE_HEADERS

    assert ws.cell(row=2, column=1).value == "244111222"
    assert ws.cell(row=2, column=2).value == "Ada Lovelace"
    assert ws.cell(row=2, column=3).value == "MATH-ADA1"
    assert ws.cell(row=2, column=4).value == 2
    assert ws.cell(row=2, column=5).value == 40
    assert ws.cell(row=3, column=2).value == "Grace Hopper"

    assert ws.protection.sheet is True
    assert ws.protection.selectUnlockedCells is False
    assert ws.protection.selectLockedCells is False
    for col in (1, 2, 3, 4):
        assert ws.cell(row=2, column=col).protection.locked is True
    assert ws.cell(row=2, column=5).protection.locked is False
    assert ws.cell(row=1, column=1).protection.locked is True
    # Style assignment order must keep total_allocation editable after alignment.
    assert ws.cell(row=3, column=5).protection.locked is False
