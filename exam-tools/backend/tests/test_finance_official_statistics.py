"""Unit tests for finance centre official statistics."""

import io
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.models import CentreStructureMode, ExamOfficialDesignation, ExaminationCentreMembershipScope
from app.schemas.examination import FinanceCentreOfficialStatisticsRow
from app.schemas.timetable import TimetableDownloadFilter
from app.services.finance_official_statistics import (
    TOTALS_ROW_ID,
    build_finance_centre_official_statistics,
    build_statistics_row,
    list_centres_for_official_statistics,
    sum_statistics_rows,
)
from app.services.finance_official_statistics_export import (
    HEADER_LABELS,
    official_statistics_export_filename,
    official_statistics_workbook_bytes,
)
from app.services.finance_school_summary import invigilator_headcount


def _official(designation: ExamOfficialDesignation, num_days: int = 1) -> MagicMock:
    off = MagicMock()
    off.designation = designation
    off.num_days = num_days
    return off


def _centre(code: str = "C001", name: str = "Test Centre") -> MagicMock:
    centre = MagicMock()
    centre.id = uuid4()
    centre.code = code
    centre.name = name
    return centre


def test_invigilator_headcount() -> None:
    officials = [
        _official(ExamOfficialDesignation.INVIGILATOR, 2),
        _official(ExamOfficialDesignation.INVIGILATOR, 3),
        _official(ExamOfficialDesignation.SUPERVISOR),
    ]
    assert invigilator_headcount(officials) == 2


def test_build_statistics_row_all_roles() -> None:
    centre = _centre()
    officials = [
        _official(ExamOfficialDesignation.INVIGILATOR, 2),
        _official(ExamOfficialDesignation.INVIGILATOR, 3),
        _official(ExamOfficialDesignation.EXTERNAL_INSPECTOR),
        _official(ExamOfficialDesignation.SUPERVISOR),
        _official(ExamOfficialDesignation.ASSISTANT_SUPERVISOR),
        _official(ExamOfficialDesignation.POLICE_OFFICER),
        _official(ExamOfficialDesignation.DEPOT_KEEPER),
    ]
    row = build_statistics_row(centre, officials, expected_invigilator_days=4)
    assert row.center_code == "C001"
    assert row.invigilator_count == 2
    assert row.invigilator_days == 5
    assert row.expected_invigilator_days == 4
    assert row.invigilator_variance == 1
    assert row.external_inspector == 1
    assert row.total_officials == 7


def test_build_statistics_row_zero_officials() -> None:
    centre = _centre("C002", "Empty Centre")
    row = build_statistics_row(centre, [], expected_invigilator_days=3)
    assert row.invigilator_count == 0
    assert row.invigilator_days == 0
    assert row.expected_invigilator_days == 3
    assert row.invigilator_variance == -3
    assert row.total_officials == 0


def test_sum_statistics_rows() -> None:
    id1, id2 = uuid4(), uuid4()
    rows = [
        FinanceCentreOfficialStatisticsRow(
            center_id=id1,
            center_code="A",
            center_name="Centre A",
            invigilator_count=2,
            invigilator_days=4,
            expected_invigilator_days=3,
            invigilator_variance=1,
            external_inspector=1,
            supervisor=1,
            total_officials=5,
        ),
        FinanceCentreOfficialStatisticsRow(
            center_id=id2,
            center_code="B",
            center_name="Centre B",
            invigilator_count=1,
            invigilator_days=2,
            expected_invigilator_days=5,
            invigilator_variance=-3,
            assistant_supervisor=1,
            depot_keeper=1,
            total_officials=3,
        ),
    ]
    totals = sum_statistics_rows(rows)
    assert totals.center_id == TOTALS_ROW_ID
    assert totals.center_code == "TOTAL"
    assert totals.invigilator_count == 3
    assert totals.invigilator_days == 6
    assert totals.expected_invigilator_days == 8
    assert totals.invigilator_variance == -2
    assert totals.total_officials == 8


def test_official_statistics_export_filename() -> None:
    name = official_statistics_export_filename("2026 MAY — BECE", TimetableDownloadFilter.CORE_ONLY)
    assert name == "2026 MAY — BECE official-statistics CORE.xlsx"


def test_official_statistics_workbook_bytes_includes_styled_headers_and_totals() -> None:
    centre_id = uuid4()
    rows = [
        FinanceCentreOfficialStatisticsRow(
            center_id=centre_id,
            center_code="C001",
            center_name="Host",
            invigilator_count=1,
            invigilator_days=2,
            expected_invigilator_days=1,
            invigilator_variance=1,
            total_officials=1,
        ),
    ]
    totals = sum_statistics_rows(rows)
    payload = official_statistics_workbook_bytes(
        rows,
        totals=totals,
        exam_label="2026 BECE",
        subject_filter=TimetableDownloadFilter.ALL,
    )
    wb = load_workbook(io.BytesIO(payload))
    ws = wb["Official statistics"]
    assert ws.cell(row=1, column=1).value == "Official statistics — 2026 BECE"
    assert ws.cell(row=4, column=1).value == HEADER_LABELS[0]
    assert ws.cell(row=5, column=1).value == "C001"
    assert ws.cell(row=5, column=6).value == 1
    assert ws.cell(row=6, column=1).value == "TOTAL"
    assert ws.cell(row=6, column=6).value == 1


@pytest.mark.asyncio
async def test_list_centres_for_official_statistics_split_core() -> None:
    exam = MagicMock()
    exam.centre_structure_mode = CentreStructureMode.SPLIT
    session = AsyncMock()
    session.get.return_value = exam
    core_centre = MagicMock()
    both_centre = MagicMock()

    with patch(
        "app.services.finance_official_statistics.list_centres_for_examination",
        new_callable=AsyncMock,
    ) as mock_list:
        mock_list.return_value = [core_centre, both_centre]
        result = await list_centres_for_official_statistics(
            session, 1, TimetableDownloadFilter.CORE_ONLY
        )

    mock_list.assert_awaited_once_with(
        session, 1, membership_scope=ExaminationCentreMembershipScope.CORE
    )
    assert result == [core_centre, both_centre]


@pytest.mark.asyncio
async def test_list_centres_for_official_statistics_all_includes_every_centre() -> None:
    exam = MagicMock()
    exam.centre_structure_mode = CentreStructureMode.SPLIT
    session = AsyncMock()
    session.get.return_value = exam
    all_centres = [MagicMock(), MagicMock(), MagicMock()]

    with patch(
        "app.services.finance_official_statistics.list_centres_for_examination",
        new_callable=AsyncMock,
    ) as mock_list:
        mock_list.return_value = all_centres
        result = await list_centres_for_official_statistics(
            session, 1, TimetableDownloadFilter.ALL
        )

    mock_list.assert_awaited_once_with(session, 1)
    assert result == all_centres


@pytest.mark.asyncio
async def test_build_finance_centre_official_statistics_excludes_non_host_centres() -> None:
    core_centre = _centre("CORE1", "Core host")
    session = AsyncMock()
    loaded_row = build_statistics_row(core_centre, [], expected_invigilator_days=2)

    with (
        patch(
            "app.services.finance_official_statistics.list_centres_for_official_statistics",
            new_callable=AsyncMock,
        ) as mock_centres,
        patch(
            "app.services.finance_official_statistics.load_officials_grouped_by_centre",
            new_callable=AsyncMock,
        ) as mock_officials,
        patch(
            "app.services.finance_official_statistics.build_statistics_row_for_centre",
            new_callable=AsyncMock,
        ) as mock_row,
    ):
        mock_centres.return_value = [core_centre]
        mock_officials.return_value = {}
        mock_row.return_value = loaded_row

        response = await build_finance_centre_official_statistics(
            session,
            1,
            TimetableDownloadFilter.CORE_ONLY,
            build_invigilator_item=AsyncMock(),
        )

    assert len(response.centres) == 1
    assert response.centres[0].center_code == "CORE1"
    assert response.centres[0].expected_invigilator_days == 2
