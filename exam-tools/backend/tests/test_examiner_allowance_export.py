"""Unit tests for examiner detail Excel export optional fields and incomplete highlighting."""

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook

from app.schemas.admin_examiner_allowance import AdminExaminerAllowanceRow
from app.schemas.examination_examiner_allowance_rate import SubjectMarkingBreakdownRow
from app.services.examiner_allowance_export import (
    detail_workbook_bytes,
    parse_include_fields,
)


def _row(
    *,
    name: str = "Alice",
    bank_name: str | None = "GCB",
    branch_name: str | None = "Kumasi",
    bank_code: str | None = "001234",
    account: str | None = "1234567890123",
    travel_zone: str | None = "Zone A",
    subject_names: str = "Mathematics",
) -> AdminExaminerAllowanceRow:
    return AdminExaminerAllowanceRow(
        id=uuid4(),
        examination_id=1,
        examination_label="BECE 2026",
        full_name=name,
        examiner_type="assistant_examiner",
        region="Ashanti",
        subject_codes="MATH",
        subject_names=subject_names,
        bank_branch_id=uuid4() if account else None,
        bank_code=bank_code,
        bank_name=bank_name,
        branch_name=branch_name,
        account_number=account,
        phone_number="0550000000",
        responsibility_allowance_ghs=Decimal("100"),
        inconvenience_allowance_ghs=Decimal("0"),
        chief_examiners_report_ghs=Decimal("0"),
        vetting_of_scripts_ghs=Decimal("0"),
        internal_commuting_ghs=Decimal("0"),
        marking_allowance_ghs=Decimal("0"),
        travel_base_ghs=Decimal("0"),
        travel_zone_name=travel_zone,
        travel_role_factor=Decimal("1"),
        travel_and_transport_ghs=Decimal("0"),
        total_allocated_scripts=0,
        marking_withholding_tax_ghs=Decimal("0"),
        marking_net_ghs=Decimal("0"),
        vetting_withholding_tax_ghs=Decimal("0"),
        vetting_net_ghs=Decimal("0"),
        payout_travel_commuting_ghs=Decimal("0"),
        payout_allowances_marking_ghs=Decimal("100"),
        total_payable_ghs=Decimal("100"),
        subject_breakdowns=[
            SubjectMarkingBreakdownRow(
                subject_id=1,
                subject_code="MATH",
                subject_name="Mathematics",
                paper_number=1,
                allocated_booklets=0,
                marking_allowance_ghs=Decimal("0"),
            )
        ],
        created_at=datetime(2026, 6, 1),
        updated_at=datetime(2026, 6, 1),
    )


def test_parse_include_fields_allowlist() -> None:
    assert parse_include_fields(None) == frozenset()
    assert parse_include_fields("travel_zone,subject_names,unknown") == frozenset(
        {"travel_zone", "subject_names"}
    )


def test_detail_workbook_optional_columns() -> None:
    payload = detail_workbook_bytes(
        [_row()],
        title="Test",
        include_fields=frozenset({"travel_zone", "subject_names"}),
    )
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws is not None
    headers = [ws.cell(row=2, column=c).value for c in range(1, 30) if ws.cell(row=2, column=c).value]
    assert "Subject names" in headers
    assert "Travel zone" in headers
    assert headers[headers.index("Subjects") + 1] == "Subject names"
    assert headers[headers.index("Travel zone") + 1] == "T & T (GHS)"
    sn_col = headers.index("Subject names") + 1
    tz_col = headers.index("Travel zone") + 1
    assert ws.cell(row=3, column=sn_col).value == "Mathematics"
    assert ws.cell(row=3, column=tz_col).value == "Zone A"


def test_detail_workbook_highlights_incomplete_bank() -> None:
    payload = detail_workbook_bytes(
        [_row(account=""), _row(name="Bob")],
        title="Test",
    )
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws is not None
    headers = [ws.cell(row=2, column=c).value for c in range(1, 30) if ws.cell(row=2, column=c).value]
    assert "Bank status" in headers
    assert "Phone" in headers
    status_col = headers.index("Bank status") + 1
    phone_col = headers.index("Phone") + 1
    assert ws.cell(row=3, column=status_col).value == "Incomplete"
    assert ws.cell(row=4, column=status_col).value == "OK"
    assert ws.cell(row=3, column=phone_col).value == "0550000000"
    # Incomplete row uses amber fill FDE68A
    assert ws.cell(row=3, column=1).fill.fgColor.rgb in ("00FDE68A", "FDE68A")
    # Complete second row uses zebra alt
    assert ws.cell(row=4, column=1).fill.fgColor.rgb in ("00F8FAFC", "F8FAFC")


def test_bog_rows_include_phone_and_incomplete_flag() -> None:
    from app.services.examiner_allowance_bog_export import bog_rows_from_admin_items

    items = [_row(name="Alice", account=""), _row(name="Bob")]
    rows = bog_rows_from_admin_items(items)
    by_name = {r.full_name: r for r in rows}
    assert by_name["ALICE"].phone_number == "0550000000"
    assert by_name["ALICE"].incomplete_bank is True
    assert by_name["BOB"].incomplete_bank is False


def test_examiner_bog_workbook_includes_phone_and_highlights_incomplete() -> None:
    from app.services.examiner_allowance_bog_export import bog_rows_from_admin_items
    from app.services.exam_official_bog_export import bog_workbook_bytes

    items = [_row(name="Alice", account=""), _row(name="Bob")]
    rows = bog_rows_from_admin_items(items)
    payload = bog_workbook_bytes([], {}, title="Test", prebuilt_rows=rows, include_phone=True)
    wb = load_workbook(BytesIO(payload))
    ws = wb.active
    assert ws is not None
    headers = [ws.cell(row=2, column=c).value for c in range(1, 10) if ws.cell(row=2, column=c).value]
    assert headers == [
        "Serial",
        "Sort code",
        "Account number",
        "Name",
        "Phone",
        "Designation",
        "Amount (GHS)",
    ]
    assert ws.cell(row=3, column=5).value == "0550000000"
    assert ws.cell(row=3, column=1).fill.fgColor.rgb in ("00FDE68A", "FDE68A")
