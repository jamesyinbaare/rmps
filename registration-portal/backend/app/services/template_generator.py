"""Utility for generating Excel template files for bulk uploads."""

import io
from typing import TYPE_CHECKING

import pandas as pd
from datetime import date, time

from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from sqlalchemy.ext.asyncio import AsyncSession


def generate_programme_template() -> bytes:
    """
    Generate Excel template for programme upload.

    Returns:
        Bytes of Excel file
    """
    data = {
        "code": ["PROG01", "PROG02"],
        "name": ["Example Programme 1", "Example Programme 2"],
    }
    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Programmes")
    output.seek(0)
    return output.getvalue()


def generate_subject_template() -> bytes:
    """
    Generate Excel template for subject upload.

    Returns:
        Bytes of Excel file
    """
    data = {
        "code": ["301", "302", "303", "701"],
        "original_code": ["MATH301", "ENG302", "PHY303", "SCI701"],
        "name": ["Mathematics", "English", "Physics", "Science"],
        "subject_type": ["CORE", "CORE", "CORE", "ELECTIVE"],
        "choice_group_id": ["", "1", "1", ""],
    }
    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Subjects")
    output.seek(0)
    return output.getvalue()


def generate_candidate_template(exam_series: str | None = None) -> bytes:
    """
    Generate Excel template for candidate bulk upload.
    All columns are formatted as text to preserve leading zeros.

    Args:
        exam_series: Exam series (e.g., "MAY/JUNE", "NOV/DEC").
        - For "MAY/JUNE": no subject columns (optional_core_groups and optional_subjects excluded)
        - For "NOV/DEC": subject_ids column is required (comma-separated subject IDs)

    Returns:
        Bytes of Excel file
    """
    # Normalize exam series
    from app.services.subject_selection import normalize_exam_series
    normalized_series = normalize_exam_series(exam_series)
    is_nov_dec = normalized_series == "NOV/DEC"

    # Required columns first
    data = {
        "firstname": ["John", "Jane"],
        "lastname": ["Doe", "Smith"],
        "othername": ["", "Mary"],
        "date_of_birth": ["2005-01-15", "2005-03-20"],
        "gender": ["M", "F"],
        "programme_code": ["PROG01", "PROG02"],
    }

    # Optional columns
    optional_cols = {
        "national_id": ["123456789", "987654321"],
        "contact_email": ["john@example.com", "jane@example.com"],
        "contact_phone": ["+1234567890", "+0987654321"],
        "address": ["123 Main St", "456 Oak Ave"],
        "guardian_name": ["John Doe Sr.", "Jane Smith Sr."],
        "guardian_phone": ["+1234567891", "+0987654322"],
        "disability": ["", "Visual"],
        "guardian_digital_address": ["GA-123-4567", "GA-789-0123"],
        "guardian_national_id": ["GHA-123456789-1", "GHA-987654321-2"],
    }

    # For NOV/DEC: registration_type is not included (always enforced as "referral" on backend)
    # For MAY/JUNE: include registration_type as optional
    if not is_nov_dec:
        optional_cols["registration_type"] = ["free_tvet", "referral"]

    data.update(optional_cols)

    # Subject columns only for NOV/DEC (explicitly check for NOV/DEC)
    # For NOV/DEC: include subject_codes column (comma-separated list of subject original codes)
    # For MAY/JUNE: no subject columns
    # For unknown/None exam_series: no subject columns (default to MAY/JUNE behavior)
    if is_nov_dec:
        data.update({
            "subject_codes": ["C701,C702", "C30-1-01"],  # Example: comma-separated subject original codes
        })

    # Convert all values to strings to ensure text formatting
    df = pd.DataFrame(data)
    for col in df.columns:
        df[col] = df[col].astype(str)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Candidates")

        # Get the worksheet to format columns as text
        worksheet = writer.sheets["Candidates"]

        # Format all columns as text and add colors to headers
        # Define colors for headers
        required_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # Light green
        optional_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Light yellow

        # Define required columns
        required_columns = {"firstname", "lastname", "date_of_birth", "gender", "programme_code"}
        # For NOV/DEC, subject_codes is also required
        if is_nov_dec:
            required_columns.add("subject_codes")

        # Get column headers from first row
        header_row = 1
        for col_idx, col in enumerate(worksheet.columns, start=1):
            header_cell = worksheet.cell(row=header_row, column=col_idx)
            column_name = str(header_cell.value).lower() if header_cell.value else ""

            # Format header cell
            header_cell.font = Font(bold=True)

            # Apply color based on whether column is required or optional
            if column_name in required_columns:
                header_cell.fill = required_fill
            else:
                header_cell.fill = optional_fill

            # Format data cells as text
            for cell in col:
                if cell.row > header_row:  # Data rows
                    cell.number_format = FORMAT_TEXT
                    # Ensure value is stored as text
                    if cell.value is not None:
                        cell.value = str(cell.value)

    output.seek(0)
    return output.getvalue()


async def generate_schedule_template(
    session: "AsyncSession",
    exam_year: int | None = None,  # noqa: ARG001
    exam_series: str | None = None,  # noqa: ARG001
    exam_type: str | None = None,  # noqa: ARG001
) -> bytes:
    """
    Generate Excel template for schedule upload prepopulated with subjects.

    Args:
        session: Database session to query subjects
        exam_year: Exam year for filename
        exam_series: Exam series for filename
        exam_type: Exam type for filename

    Returns:
        Bytes of Excel file
    """
    from sqlalchemy import select
    from app.models import Subject

    # Query all subjects from the database
    stmt = select(Subject).order_by(Subject.code)
    result = await session.execute(stmt)
    subjects = result.scalars().all()

    # Prepare data with subjects from database
    # Note: Header shows "subject_code" for user clarity, but contains original_code values
    if subjects:
        data = {
            "subject_code": [
                subject.original_code if subject.original_code else subject.code for subject in subjects
            ],
            "subject_name": [subject.name for subject in subjects],
            "paper1_date": [""] * len(subjects),
            "paper1_start_time": [""] * len(subjects),
            "paper1_end_time": [""] * len(subjects),
            "paper2_date": [""] * len(subjects),
            "paper2_start_time": [""] * len(subjects),
            "paper2_end_time": [""] * len(subjects),
            "write_together": [0] * len(subjects),
        }
    else:
        # If no subjects, create empty template with just headers
        data = {
            "subject_code": [],
            "subject_name": [],
            "paper1_date": [],
            "paper1_start_time": [],
            "paper1_end_time": [],
            "paper2_date": [],
            "paper2_start_time": [],
            "paper2_end_time": [],
            "write_together": [],
        }

    df = pd.DataFrame(data)

    # Sample data for the second sheet
    sample_date = date.today().replace(month=1, day=15)  # Example: January 15
    sample_data = {
        "subject_code": ["C701", "C702"],
        "subject_name": ["English Language", "Mathematics"],
        "paper1_date": [sample_date, sample_date],
        "paper1_start_time": [time(9, 0), time(9, 0)],
        "paper1_end_time": [time(11, 0), time(11, 30)],
        "paper2_date": [sample_date, sample_date.replace(day=16)],
        "paper2_start_time": [time(9, 0), time(14, 0)],
        "paper2_end_time": [time(11, 0), time(16, 0)],
        "write_together": [1, 0],
    }
    df_sample = pd.DataFrame(sample_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write main template sheet
        df.to_excel(writer, index=False, sheet_name="Schedules")
        worksheet = writer.sheets["Schedules"]

        # Find column indices dynamically
        date_columns = ["paper1_date", "paper2_date"]
        time_columns = ["paper1_start_time", "paper1_end_time", "paper2_start_time", "paper2_end_time"]
        number_columns = ["write_together"]

        # Get column letters for each column type
        date_col_letters = []
        time_col_letters = []
        number_col_letters = []

        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            if col_name in date_columns:
                date_col_letters.append(col_letter)
            elif col_name in time_columns:
                time_col_letters.append(col_letter)
            elif col_name in number_columns:
                number_col_letters.append(col_letter)

        # Format date columns
        for col_letter in date_col_letters:
            for cell in worksheet[col_letter][1:]:  # Skip header
                cell.number_format = "yyyy-mm-dd"

        # Format time columns - use HH:mm for 24-hour format with leading zeros
        for col_letter in time_col_letters:
            for cell in worksheet[col_letter][1:]:  # Skip header
                cell.number_format = "HH:mm"

        # Format number columns
        for col_letter in number_col_letters:
            for cell in worksheet[col_letter][1:]:  # Skip header
                cell.number_format = "0"

        # Write sample data sheet
        df_sample.to_excel(writer, index=False, sheet_name="Sample Data")
        worksheet_sample = writer.sheets["Sample Data"]

        # Format sample data columns with same formatting
        for col_letter in date_col_letters:
            for cell in worksheet_sample[col_letter][1:]:  # Skip header
                cell.number_format = "yyyy-mm-dd"

        for col_letter in time_col_letters:
            for cell in worksheet_sample[col_letter][1:]:  # Skip header
                cell.number_format = "HH:mm"

        for col_letter in number_col_letters:
            for cell in worksheet_sample[col_letter][1:]:  # Skip header
                cell.number_format = "0"

    output.seek(0)
    return output.getvalue()


async def generate_subject_pricing_template(session: "AsyncSession", exam_id: int | None = None) -> bytes:
    """
    Generate Excel template for subject pricing upload prepopulated with subjects.

    Args:
        session: Database session to query subjects
        exam_id: Optional exam ID to include existing pricing

    Returns:
        Bytes of Excel file
    """
    from sqlalchemy import select
    from app.models import Subject, SubjectPricing

    # Query all subjects from the database
    stmt = select(Subject).order_by(Subject.code)
    result = await session.execute(stmt)
    subjects = result.scalars().all()

    # Get existing pricing if exam_id is provided
    existing_pricing = {}
    if exam_id:
        pricing_stmt = select(SubjectPricing).where(
            SubjectPricing.exam_id == exam_id,
            SubjectPricing.is_active == True
        )
        pricing_result = await session.execute(pricing_stmt)
        for pricing in pricing_result.scalars().all():
            existing_pricing[pricing.subject_id] = pricing.price

    # Prepare data with subjects from database
    if subjects:
        data = {
            "original_code": [
                subject.original_code if subject.original_code else subject.code for subject in subjects
            ],
            "subject_name": [subject.name for subject in subjects],
            "price": [
                str(existing_pricing.get(subject.id, "")) for subject in subjects
            ],
        }
    else:
        # If no subjects, create empty template with just headers
        data = {
            "original_code": [],
            "subject_name": [],
            "price": [],
        }

    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="SubjectPricing")
    output.seek(0)
    return output.getvalue()


async def generate_programme_pricing_template(session: "AsyncSession", exam_id: int | None = None) -> bytes:
    """
    Generate Excel template for programme pricing upload prepopulated with programmes.

    Args:
        session: Database session to query programmes
        exam_id: Optional exam ID to include existing pricing

    Returns:
        Bytes of Excel file
    """
    from sqlalchemy import select
    from app.models import Programme, ProgrammePricing

    # Query all programmes from the database
    stmt = select(Programme).order_by(Programme.code)
    result = await session.execute(stmt)
    programmes = result.scalars().all()

    # Get existing pricing if exam_id is provided
    existing_pricing = {}
    if exam_id:
        pricing_stmt = select(ProgrammePricing).where(
            ProgrammePricing.exam_id == exam_id,
            ProgrammePricing.is_active == True
        )
        pricing_result = await session.execute(pricing_stmt)
        for pricing in pricing_result.scalars().all():
            existing_pricing[pricing.programme_id] = pricing.price

    # Prepare data with programmes from database
    if programmes:
        data = {
            "programme_code": [programme.code for programme in programmes],
            "programme_name": [programme.name for programme in programmes],
            "price": [
                str(existing_pricing.get(programme.id, "")) for programme in programmes
            ],
        }
    else:
        # If no programmes, create empty template with just headers
        data = {
            "programme_code": [],
            "programme_name": [],
            "price": [],
        }

    df = pd.DataFrame(data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="ProgrammePricing")
    output.seek(0)
    return output.getvalue()
